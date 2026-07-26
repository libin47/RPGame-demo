"""
生成事件.xlsx 和 效果.xlsx 的测试数据
严格遵循 types/event.ts 和 types/effect.ts 中定义的数据结构
优先使用场景测试数据中已关联的事件ID。
必须保持原有文件样式和格式不变。
"""

import os
import sys
import io
import copy
from copy import copy as copy_cell

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    print("错误：需要 openpyxl 库。请运行: pip install openpyxl")
    sys.exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EVENT_XLSX = os.path.join(SCRIPT_DIR, "事件.xlsx")
EFFECT_XLSX = os.path.join(SCRIPT_DIR, "效果.xlsx")


# ============================================================
# 工具函数
# ============================================================

def ensure_sheet_order(wb, sheet_names: list[str]):
    """确保工作表顺序一致"""
    for i, name in enumerate(sheet_names):
        if name in wb.sheetnames:
            current_idx = wb.sheetnames.index(name)
            if current_idx != i:
                wb.move_sheet(name, offset=i - current_idx)


def copy_cell_style(src_cell, dst_cell):
    """复制单元格样式"""
    if src_cell.has_style:
        dst_cell.font = copy_cell(src_cell.font)
        dst_cell.fill = copy_cell(src_cell.fill)
        dst_cell.border = copy_cell(src_cell.border)
        dst_cell.alignment = copy_cell(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy_cell(src_cell.protection)


def insert_data_rows(ws, start_row: int, data_rows: list[list]):
    """
    在工作表的指定行处插入数据，将后续行下移。
    data_rows: 行列表，每行是一个值列表
    """
    row_count = len(data_rows)
    if row_count == 0:
        return

    # 将现有行下移
    ws.insert_rows(start_row, row_count)

    # 写入数据
    for r_idx, row_data in enumerate(data_rows):
        row_num = start_row + r_idx
        for c_idx, val in enumerate(row_data):
            cell = ws.cell(row=row_num, column=c_idx + 1, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")


# ============================================================
# 测试数据定义
# ============================================================

# 场景测试数据中已关联的事件ID列表
# event_plane_wreckage, event_beach_crab_encounter, event_strange_trees,
# event_gather_berries, event_gather_wood, event_glowing_moss,
# event_journal_fragment, event_cave_markings

EVENT_DATA = {
    # ===== 飞机残骸事件 =====
    "event_plane_wreckage": {
        "name": "搜索飞机残骸",
        "eventType": "NORMAL",
        "frames": [
            {
                "id": "wreckage_search",
                "order": 1,
                "text": "你穿过敞开的变形舱门，看到了扭曲的金属和碎裂的行李箱。",
                "options": [
                    {
                        "id": "find_in_cloth",
                        "text": "翻找残骸",
                        "displayPriority": 3,
                        "isOneTime": "True",
                        "selectedFlag": "selected_find_in_cloth",
                        "result_type": "nextFrame",
                        "result_target": "after_gather",
                        "result_text": "你在残骸中找到了半瓶矿泉水、压缩饼干、绷带和一把多功能刀。",
                        "result_effects": "eff_plane_add_cloth,eff_plane_add_knife",
                    },
                    {
                        "id": "take_sword",
                        "text": "拿起生锈的铁剑",
                        "displayPriority": 2,
                        "isOneTime": "True",
                        "selectedFlag": "selected_take_sword",
                        "result_type": "nextFrame",
                        "result_target": "after_gather",
                        "result_text": "你从座椅下抽出了一把生锈的铁剑。",
                        "result_effects": "eff_plane_add_sword",
                        "result_setFlags": "collected_sword:true",
                    },
                ],
            },
            {
                "id": "after_gather",
                "order": 2,
                "text": "你收好了搜刮到的物资。当前最重要的是找个安全的地方度过第一夜。",
                "options": [
                    {
                        "id": "leave_wreckage",
                        "text": "离开残骸",
                        "displayPriority": 1,
                        "result_type": "endEvent",
                        "result_exitText": "你离开了飞机残骸，回到了海滩上",
                    },
                ],
            },
        ],
    },

    # ===== 变异蟹遭遇事件 =====
    "event_beach_crab_encounter": {
        "name": "海滩遇蟹",
        "eventType": "BATTLE",
        "isRepeatable": "True",
        "frames": [
            {
                "id": "crab_spotted",
                "order": 1,
                "text": "一只体型巨大的变异蟹从沙中钻出，挥舞着巨大的螯钳向你逼近。",
                "options": [
                    {
                        "id": "fight_crab",
                        "text": "战斗",
                        "displayPriority": 2,
                        "optionStyle": "DANGER",
                        "result_type": "triggerBattle",
                        "result_enemyId": "mutated_crab",
                        "result_victoryFrameId": "crab_victory",
                        "result_defeatFrameId": "crab_defeat",
                        "result_escapeFrameId": "crab_escaped",
                        "result_canEscape": "True",
                        "result_firstEncounterBonus": "True",
                    },
                    {
                        "id": "flee_crab",
                        "text": "逃跑",
                        "displayPriority": 1,
                        "description": "尝试逃离变异蟹",
                        "costs": "cost_crab_flee",
                        "result_type": "endEvent",
                        "result_exitText": "你飞快地逃离了变异蟹的领地",
                    },
                ],
            },
            {
                "id": "crab_victory",
                "order": 2,
                "text": "变异蟹轰然倒地，不再动弹。",
                "onEnterEffects": "eff_crab_victory_flag,eff_crab_victory_exp",
                "options": [
                    {
                        "id": "butcher_crab",
                        "text": "分解蟹肉",
                        "displayPriority": 1,
                        "result_type": "endEvent",
                        "result_exitText": "你从变异蟹身上获取了一些有用的材料",
                    },
                ],
            },
            {
                "id": "crab_escaped",
                "order": 2,
                "text": "你抓住机会逃离了战斗。变异蟹没有追上来。",
                "options": [
                    {
                        "id": "return_beach",
                        "text": "返回海滩",
                        "displayPriority": 1,
                        "result_type": "endEvent",
                        "result_exitText": "你安全回到了海滩上",
                    },
                ],
            },
            {
                "id": "crab_defeat",
                "order": 2,
                "text": "变异蟹的巨螯击中了你的要害，你失去了意识...",
                "options": [
                    {
                        "id": "accept_defeat",
                        "text": "...",
                        "displayPriority": 1,
                        "result_type": "endEvent",
                        "result_exitText": "",
                    },
                ],
            },
        ],
    },

    # ===== 变异树木事件 =====
    "event_strange_trees": {
        "name": "变异树木",
        "eventType": "NORMAL",
        "frames": [
            {
                "id": "approach_trees",
                "order": 1,
                "text": "你走近那些树干上有异常突起的树木。树皮表面覆盖着一种暗紫色的苔藓，散发着微弱的光芒。",
                "options": [
                    {
                        "id": "touch_moss",
                        "text": "触摸苔藓",
                        "displayPriority": 2,
                        "result_type": "nextFrame",
                        "result_target": "touch_result",
                        "result_text": "你伸手触碰了那暗紫色的苔藓。\n触感冰凉而湿润，一种奇异的共鸣感顺着指尖传遍全身。",
                        "result_effects": "eff_strange_moss_san",
                    },
                    {
                        "id": "leave_trees",
                        "text": "离开",
                        "displayPriority": 1,
                        "result_type": "endEvent",
                        "result_exitText": "你决定不去碰那些可疑的树木",
                    },
                ],
            },
            {
                "id": "touch_result",
                "order": 2,
                "text": "你感到一阵眩晕。那些苔藓在你的注视下似乎在微微蠕动。\n也许这不是什么好兆头。",
                "options": [
                    {
                        "id": "back_away",
                        "text": "后退离开",
                        "displayPriority": 1,
                        "result_type": "endEvent",
                        "result_exitText": "你快步离开了这片区域，但那股冰凉的感觉仍停留在指尖",
                    },
                ],
            },
        ],
    },

    # ===== 采集浆果事件 =====
    "event_gather_berries": {
        "name": "采集浆果",
        "eventType": "NORMAL",
        "isRepeatable": "True",
        "frames": [
            {
                "id": "gather_start",
                "order": 1,
                "text": "灌木丛中长满了红色的浆果，看起来可以食用。你要采集一些吗？",
                "options": [
                    {
                        "id": "gather_berries",
                        "text": "采集浆果",
                        "displayPriority": 1,
                        "costs": "cost_gather_berries",
                        "result_type": "nextFrame",
                        "result_target": "gather_result",
                        "result_effects": "eff_gather_exp",
                    },
                    {
                        "id": "leave_berries",
                        "text": "离开",
                        "displayPriority": 2,
                        "result_type": "endEvent",
                        "result_exitText": "你决定不采摘这些浆果",
                    },
                ],
            },
            {
                "id": "gather_result",
                "order": 2,
                "text": "你小心地采摘了一些浆果。",
                "onEnterEffects": "eff_gather_berry_item",
                "options": [
                    {
                        "id": "done_gathering",
                        "text": "继续前进",
                        "displayPriority": 1,
                        "result_type": "endEvent",
                        "result_exitText": "你继续探索森林",
                    },
                ],
            },
        ],
    },

    # ===== 收集木材事件 =====
    "event_gather_wood": {
        "name": "收集落枝",
        "eventType": "NORMAL",
        "isRepeatable": "True",
        "frames": [
            {
                "id": "gather_wood_start",
                "order": 1,
                "text": "地面上散落着不少干枯的树枝，是生火的好材料。",
                "options": [
                    {
                        "id": "collect_wood",
                        "text": "收集树枝",
                        "displayPriority": 1,
                        "costs": "cost_gather_wood",
                        "result_type": "nextFrame",
                        "result_target": "wood_result",
                        "result_effects": "eff_wood_collect",
                    },
                    {
                        "id": "ignore_wood",
                        "text": "忽略",
                        "displayPriority": 2,
                        "result_type": "endEvent",
                        "result_exitText": "你决定不收集这些树枝",
                    },
                ],
            },
            {
                "id": "wood_result",
                "order": 2,
                "text": "你收集了一些干树枝，可以作为引火材料使用。",
                "options": [
                    {
                        "id": "continue_explore",
                        "text": "继续探索",
                        "displayPriority": 1,
                        "result_type": "endEvent",
                        "result_exitText": "你继续在森林中探索",
                    },
                ],
            },
        ],
    },

    # ===== 发光苔藓事件 =====
    "event_glowing_moss": {
        "name": "发光苔藓",
        "eventType": "NORMAL",
        "isRepeatable": "True",
        "frames": [
            {
                "id": "observe_moss",
                "order": 1,
                "text": "洞穴墙壁上覆盖着一层发着幽蓝光芒的苔藓，照亮了周围一小片区域。",
                "options": [
                    {
                        "id": "collect_moss",
                        "text": "采集苔藓",
                        "displayPriority": 1,
                        "result_type": "nextFrame",
                        "result_target": "moss_collected",
                        "result_text": "你小心地刮下了一些发光苔藓，它们在你手中微弱地闪烁着。",
                        "result_effects": "eff_moss_collect",
                    },
                    {
                        "id": "leave_moss",
                        "text": "忽略",
                        "displayPriority": 2,
                        "result_type": "endEvent",
                        "result_exitText": "你决定不去碰那些发光的苔藓",
                    },
                ],
            },
            {
                "id": "moss_collected",
                "order": 2,
                "text": "你获得了少量的发光苔藓。它们也许能在黑暗中提供照明。",
                "options": [
                    {
                        "id": "continue_cave",
                        "text": "继续探索洞穴",
                        "displayPriority": 1,
                        "result_type": "endEvent",
                        "result_exitText": "你继续在洞穴中探索",
                    },
                ],
            },
        ],
    },

    # ===== 研究日志事件 =====
    "event_journal_fragment": {
        "name": "发现研究日志",
        "eventType": "NORMAL",
        "isRepeatable": "False",
        "triggeredFlag": "triggered_journal_fragment",
        "frames": [
            {
                "id": "find_journal",
                "order": 1,
                "text": "你在洞穴的角落发现了几页发黄的纸，上面密密麻麻地写着字。",
                "options": [
                    {
                        "id": "read_journal",
                        "text": "阅读日志",
                        "displayPriority": 1,
                        "result_type": "nextFrame",
                        "result_target": "read_journal_content",
                        "result_effects": "eff_journal_add_item,eff_journal_flag",
                    },
                    {
                        "id": "ignore_journal",
                        "text": "不理会",
                        "displayPriority": 2,
                        "result_type": "endEvent",
                        "result_exitText": "你决定不去碰那些可疑的纸张",
                    },
                ],
            },
            {
                "id": "read_journal_content",
                "order": 2,
                "text": "你翻开日志，上面的内容让你不寒而栗。\n日志记载着关于某种孢子的研究，内容令人不安。",
                "onEnterEffects": "eff_journal_san_loss",
                "options": [
                    {
                        "id": "close_journal",
                        "text": "合上日志",
                        "displayPriority": 1,
                        "result_type": "endEvent",
                        "result_exitText": "你合上了日志，但那些文字仍然在你脑海中回荡",
                    },
                ],
            },
        ],
    },

    # ===== 洞穴刻痕事件 =====
    "event_cave_markings": {
        "name": "奇怪的刻痕",
        "eventType": "NORMAL",
        "frames": [
            {
                "id": "examine_markings",
                "order": 1,
                "text": "洞穴墙壁上有一些奇怪的刻痕，看起来不像是自然形成的。\n这些符号排列有序，似乎蕴含着某种意义。",
                "options": [
                    {
                        "id": "study_markings",
                        "text": "仔细研究",
                        "displayPriority": 1,
                        "result_type": "nextFrame",
                        "result_target": "markings_studied",
                        "result_text": "你花了一些时间端详这些刻痕。它们似乎是在描述某种祭祀仪式。\n你的SAN值下降了。",
                        "result_effects": "eff_markings_san",
                    },
                    {
                        "id": "ignore_markings",
                        "text": "不去理会",
                        "displayPriority": 2,
                        "result_type": "endEvent",
                        "result_exitText": "你决定不去深究这些诡异的符号",
                    },
                ],
            },
            {
                "id": "markings_studied",
                "order": 2,
                "text": "那些符号深深刻在了你的脑海中。你预感这座岛屿的秘密远比你想象的要深。",
                "options": [
                    {
                        "id": "leave_cave_area",
                        "text": "离开",
                        "displayPriority": 1,
                        "result_type": "endEvent",
                        "result_exitText": "你离开了那片刻有符号的墙壁",
                    },
                ],
            },
        ],
    },
}

# 效果数据定义
# ID列表: 效果结果ID -> 效果ID的映射，用于在效果.xlsx中创建数据
EFFECT_RESULTS_DATA = {
    # 飞机残骸事件效果
    "eff_plane_add_cloth": {"effect_id": "effect_plane_cloth", "probability": 1, "description": "获得破布"},
    "eff_plane_add_knife": {"effect_id": "effect_plane_knife", "probability": 1, "description": "获得多功能刀"},
    "eff_plane_add_sword": {"effect_id": "effect_plane_sword", "probability": 1, "description": "获得生锈的铁剑"},
    # 变异蟹事件效果
    "eff_crab_victory_flag": {"effect_id": "effect_crab_flag", "probability": 1, "description": "设置击败变异蟹标志"},
    "eff_crab_victory_exp": {"effect_id": "effect_crab_exp", "probability": 1, "description": "获得探索经验"},
    # 变异树木事件效果
    "eff_strange_moss_san": {"effect_id": "effect_moss_san", "probability": 1, "description": "SAN-5"},
    # 采集浆果事件效果
    "eff_gather_exp": {"effect_id": "effect_gather_exp", "probability": 1, "description": "获得采集经验"},
    "eff_gather_berry_item": {"effect_id": "effect_gather_berry", "probability": 1, "description": "获得浆果"},
    # 收集木材事件效果
    "eff_wood_collect": {"effect_id": "effect_wood_add", "probability": 1, "description": "获得木材"},
    # 发光苔藓事件效果
    "eff_moss_collect": {"effect_id": "effect_moss_item", "probability": 1, "description": "获得发光苔藓"},
    # 研究日志事件效果
    "eff_journal_add_item": {"effect_id": "effect_journal_item", "probability": 1, "description": "获得日志碎片"},
    "eff_journal_flag": {"effect_id": "effect_journal_flag", "probability": 1, "description": "设置已找到日志碎片"},
    "eff_journal_san_loss": {"effect_id": "effect_journal_san", "probability": 1, "description": "SAN-10"},
    # 洞穴刻痕事件效果
    "eff_markings_san": {"effect_id": "effect_markings_san", "probability": 1, "description": "SAN-5"},
}

# 效果本身的数据
EFFECTS_DATA = {
    "effect_plane_cloth": {"type": "ITEM", "itemId": "cloth_scrap", "changeType": "ADD", "quantity": 3},
    "effect_plane_knife": {"type": "ITEM", "itemId": "multitool_knife", "changeType": "ADD", "quantity": 1},
    "effect_plane_sword": {"type": "ITEM", "itemId": "rusty_sword", "changeType": "ADD", "quantity": 1},
    "effect_crab_flag": {"type": "FLAG", "flagId": "defeated_first_crab", "operation": "SET", "value": "true"},
    "effect_crab_exp": {"type": "GAIN_EXP", "target": "SURVIVAL_SKILL", "targetId": "exploration", "amount": 30},
    "effect_moss_san": {"type": "ATTRIBUTE", "attribute": "SAN", "operation": "SUBTRACT", "value": 5},
    "effect_gather_exp": {"type": "GAIN_EXP", "target": "SURVIVAL_SKILL", "targetId": "gathering", "amount": 15},
    "effect_gather_berry": {"type": "ITEM", "itemId": "wild_berries", "changeType": "ADD", "quantity": 3},
    "effect_wood_add": {"type": "ITEM", "itemId": "firewood", "changeType": "ADD", "quantity": 5},
    "effect_moss_item": {"type": "ITEM", "itemId": "glowing_moss", "changeType": "ADD", "quantity": 2},
    "effect_journal_item": {"type": "ITEM", "itemId": "journal_fragment", "changeType": "ADD", "quantity": 1},
    "effect_journal_flag": {"type": "FLAG", "flagId": "found_journal_fragment", "operation": "SET", "value": "true"},
    "effect_journal_san": {"type": "ATTRIBUTE", "attribute": "SAN", "operation": "SUBTRACT", "value": 10},
    "effect_markings_san": {"type": "ATTRIBUTE", "attribute": "SAN", "operation": "SUBTRACT", "value": 5},
}

# 选项消耗数据
EVENT_COSTS_DATA = [
    {"id": "cost_crab_flee", "costType": "STAMINA", "value": 10},
    {"id": "cost_gather_berries", "costType": "STAMINA", "value": 8},
    {"id": "cost_gather_wood", "costType": "STAMINA", "value": 5},
]


# ============================================================
# 向事件.xlsx写入测试数据
# ============================================================

def populate_event_xlsx():
    print("▶ 正在填充事件.xlsx测试数据...")
    wb = openpyxl.load_workbook(EVENT_XLSX)

    # ---- 事件 Sheet ----
    ws = wb["事件"]
    # 需要插入的数据行（基于事件字典构建）
    event_rows = []
    for evt_id, evt in EVENT_DATA.items():
        frames_str = ",".join(f['id'] for f in evt['frames'])
        on_enter_str = ""
        if evt.get('frames') and evt['frames'][0].get('onEnterEffects'):
            on_enter_str = evt['frames'][0]['onEnterEffects']
        event_rows.append([
            evt_id,
            evt['name'],
            frames_str,
            on_enter_str,
            f"EventType.{evt['eventType']}",
            "True" if evt.get('isRepeatable') else ("False" if 'isRepeatable' in evt else ""),
            evt.get('triggeredFlag', ""),
            evt.get('untriggerableText', ""),
            None,
        ])
    insert_data_rows(ws, 2, event_rows)
    print(f"  [OK] 事件: 添加了 {len(event_rows)} 条数据")

    # ---- 事件帧 Sheet ----
    ws = wb["事件帧"]
    frame_rows = []
    for evt_id, evt in EVENT_DATA.items():
        for frame in evt['frames']:
            # options: 逗号分隔的选项ID
            opt_ids = ",".join(o['id'] for o in frame['options'])
            on_enter = frame.get('onEnterEffects', "")
            frame_rows.append([
                frame['id'],
                frame['order'],
                frame['text'],
                opt_ids,
                frame.get('displayCondition', ""),
                on_enter,
                frame.get('onExitEffects', ""),
            ])
    insert_data_rows(ws, 2, frame_rows)
    print(f"  [OK] 事件帧: 添加了 {len(frame_rows)} 条数据")

    # ---- 事件选项 Sheet ----
    ws = wb["事件选项"]
    option_rows = []
    # 先收集所有选项到一个有序列表（每个选项对应一行）
    row_index = 0
    for evt_id, evt in EVENT_DATA.items():
        for frame in evt['frames']:
            for opt in frame['options']:
                row_index += 1
                result_id = f"result_{opt['id']}"
                costs = opt.get('costs', "")
                option_rows.append([
                    opt['id'],
                    opt['text'],
                    opt.get('description', ""),
                    opt.get('displayCondition', ""),
                    opt.get('availableCondition', ""),
                    opt.get('unavailableTooltip', ""),
                    costs,
                    result_id,
                    opt.get('displayPriority', 1),
                    opt.get('requiresConfirmation', ""),
                    opt.get('confirmationText', ""),
                    opt.get('isOneTime', "False"),
                    opt.get('selectedFlag', ""),
                    opt.get('textAfterSelected', ""),
                ])
    insert_data_rows(ws, 2, option_rows)
    print(f"  [OK] 事件选项: 添加了 {len(option_rows)} 条数据")

    # ---- 选项结果 Sheet ----
    ws = wb["选项结果"]
    result_rows = []
    for evt_id, evt in EVENT_DATA.items():
        for frame in evt['frames']:
            for opt in frame['options']:
                result_id = f"result_{opt['id']}"
                rtype = opt['result_type']
                effects = opt.get('result_effects', "")
                set_flags = opt.get('result_setFlags', "")

                # 根据类型构建字段
                row = [result_id, rtype, "", "", effects, set_flags, "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", None, None, None]
                row[0] = result_id
                row[1] = rtype
                row[2] = ""  # weight
                row[3] = ""  # condition
                row[4] = effects
                row[5] = set_flags

                if rtype == "nextFrame":
                    row[6] = opt.get('result_target', "")
                    row[7] = opt.get('result_text', "")
                elif rtype == "endEvent":
                    row[8] = opt.get('result_exitText', "")
                elif rtype == "triggerBattle":
                    row[9] = opt.get('result_enemyId', "")
                    row[10] = opt.get('result_victoryFrameId', "")
                    row[11] = opt.get('result_defeatFrameId', "")
                    row[12] = opt.get('result_escapeFrameId', "")
                    row[13] = "True" if opt.get('result_canEscape') else ""
                    row[14] = "True" if opt.get('result_firstEncounterBonus') else ""

                result_rows.append(row)
    insert_data_rows(ws, 2, result_rows)
    print(f"  [OK] 选项结果: 添加了 {len(result_rows)} 条数据")

    # ---- 交互花费 Sheet ----
    ws = wb["交互花费"]
    cost_rows = []
    for cost in EVENT_COSTS_DATA:
        cost_rows.append([
            cost['id'],
            f"EventOptionCostType.{cost['costType']}",
            cost['value'],
            cost.get('itemId', ""),
            cost.get('itemQuantity', ""),
            None,
            None,
        ])
    insert_data_rows(ws, 2, cost_rows)
    print(f"  [OK] 交互花费: 添加了 {len(cost_rows)} 条数据")

    wb.save(EVENT_XLSX)
    print("  [OK] 事件.xlsx 保存成功")


# ============================================================
# 向效果.xlsx写入测试数据
# ============================================================

def populate_effect_xlsx():
    print("\n▶ 正在填充效果.xlsx测试数据...")
    wb = openpyxl.load_workbook(EFFECT_XLSX)

    # ---- 效果结果 Sheet ----
    ws = wb["效果结果"]
    result_rows = []
    for eff_id, eff in EFFECT_RESULTS_DATA.items():
        result_rows.append([
            eff_id,
            eff['effect_id'],
            eff['probability'],
            eff.get('condition', ""),
            eff.get('description', ""),
        ])
    insert_data_rows(ws, 2, result_rows)
    print(f"  [OK] 效果结果: 添加了 {len(result_rows)} 条数据")

    # ---- 效果 Sheet ----
    ws = wb["效果"]
    effect_rows = []
    for eff_id, eff_data in EFFECTS_DATA.items():
        eff_type = eff_data['type']
        row = [eff_id, eff_type, "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""]
        # 44列，索引0-43
        row[0] = eff_id
        row[1] = eff_type

        if eff_type == "ATTRIBUTE":
            row[2] = f"AttributeType.{eff_data['attribute']}"
            row[3] = f"AttributeOperation.{eff_data['operation']}"
            row[4] = eff_data['value']
        elif eff_type == "ITEM":
            row[12] = eff_data['itemId']
            row[13] = f"ItemChangeType.{eff_data['changeType']}"
            row[14] = eff_data['quantity']
        elif eff_type == "FLAG":
            row[21] = eff_data['flagId']
            row[22] = f"FlagOperation.{eff_data['operation']}"
            row[23] = eff_data['value']
        elif eff_type == "GAIN_EXP":
            row[38] = f"GainExpTarget.{eff_data['target']}"
            row[39] = eff_data['targetId']
            row[40] = eff_data['amount']

        effect_rows.append(row)
    insert_data_rows(ws, 2, effect_rows)
    print(f"  [OK] 效果: 添加了 {len(effect_rows)} 条数据")

    # 确保工作表顺序
    ensure_sheet_order(wb, ["效果结果", "效果", "随机数量", "参数表", "筛选表格"])

    wb.save(EFFECT_XLSX)
    print("  [OK] 效果.xlsx 保存成功")


# ============================================================
# 主程序
# ============================================================

def main():
    print("=" * 60)
    print("事件/效果测试数据生成器")
    print("=" * 60)

    if not os.path.exists(EVENT_XLSX):
        print(f"错误: 找不到 {EVENT_XLSX}")
        sys.exit(1)
    if not os.path.exists(EFFECT_XLSX):
        print(f"错误: 找不到 {EFFECT_XLSX}")
        sys.exit(1)

    populate_event_xlsx()
    populate_effect_xlsx()

    print("\n" + "=" * 60)
    print("测试数据生成完成！")
    print("=" * 60)


if __name__ == "__main__":
    main()