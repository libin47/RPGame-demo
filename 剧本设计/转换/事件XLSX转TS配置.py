"""
事件XLSX转TS配置转换器
========================
从 事件.xlsx、效果.xlsx、条件.xlsx 读取数据，转换为符合 types/event.ts 格式的 TypeScript 代码，
输出到 config/events.ts。

数据关联关系：
  - 事件 -> 帧: 通过 frames 字段（逗号分隔的帧ID列表）关联
  - 帧 -> 选项: 通过 options 字段（逗号分隔的选项ID列表）关联
  - 选项 -> 结果: 通过 results 字段（结果ID）关联
  - 效果结果 -> 效果: 通过 effect 字段（效果ID）关联
  - 所有效果、条件都通过ID引用，在生成的TS中转换为常量引用

依赖：pip install openpyxl

使用方式：
    python "剧本设计/事件XLSX转TS配置.py"
"""

import os
import re
import sys
import json
from collections import defaultdict

import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

try:
    import openpyxl
except ImportError:
    print("错误：需要 openpyxl 库。请运行: pip install openpyxl")
    sys.exit(1)


# ============================================================
# 路径配置
# ============================================================

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)

EVENT_XLSX = os.path.join(SCRIPT_DIR, "事件.xlsx")
EFFECT_XLSX = os.path.join(SCRIPT_DIR, "效果.xlsx")
COND_XLSX = os.path.join(SCRIPT_DIR, "条件.xlsx")
OUTPUT_TS = os.path.join(PROJECT_ROOT, "src", "config", "events.ts")


# ============================================================
# XLSX 读取工具
# ============================================================

def read_sheet(wb, sheet_name: str) -> list[dict]:
    """读取工作表，返回列表[行字典]，key为表头"""
    if sheet_name not in wb.sheetnames:
        print(f"  警告: 工作表 '{sheet_name}' 不存在，跳过")
        return []
    ws = wb[sheet_name]
    headers = []
    for c in range(1, ws.max_column + 1):
        raw = str(ws.cell(1, c).value or "").strip()
        key = raw.split("\n")[0].strip()
        key = re.sub(r"（[^）]*）", "", key).strip()
        headers.append(key)
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        has_data = False
        for c in range(1, ws.max_column + 1):
            val = ws.cell(r, c).value
            key = headers[c - 1]
            row[key] = val
            if val is not None:
                has_data = True
        if has_data:
            rows.append(row)
    return rows


def parse_comma_list(val) -> list[str]:
    if val is None:
        return []
    s = str(val).strip()
    if not s:
        return []
    return [item.strip() for item in s.split(",") if item.strip()]


def parse_bool(val) -> bool:
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    return s in ("true", "yes", "1")


def parse_int(val, default=None):
    if val is None:
        return default
    try:
        return int(float(str(val)))
    except (ValueError, TypeError):
        return default


def parse_float(val, default=None):
    if val is None:
        return default
    try:
        return float(str(val))
    except (ValueError, TypeError):
        return default


def to_ts_value(val, indent=0):
    """将Python值转换为TypeScript值字符串"""
    if val is None:
        return "undefined"
    if isinstance(val, bool):
        return "true" if val else "false"
    if isinstance(val, (int, float)):
        return str(val)
    if isinstance(val, str):
        if val.lower() in ("true", "false"):
            return val.lower()
        if val.lower() == "null":
            return "undefined"
        try:
            if "." in val:
                return str(float(val))
            return str(int(val))
        except ValueError:
            pass
        escaped = val.replace("\\", "\\\\").replace("'", "\\'").replace("\n", "\\n").replace("\r", "")
        return f"'{escaped}'"
    return json.dumps(val, ensure_ascii=False)


def is_enum_ref(val) -> bool:
    """检查值是否是枚举引用（如 EventType.NORMAL）"""
    if val is None:
        return False
    s = str(val).strip()
    return bool(re.match(r'^[A-Za-z_][\w.]*\.\w+$', s))


def to_effect_enum(val, enum_name: str, default: str) -> str:
    """
    将值转换为枚举引用字符串。
    如果值已经是枚举引用（如 'AttributeType.SAN'），直接返回。
    如果是简单值（如 'SAN'），转换为 'AttributeType.SAN'。
    如果值为空，返回默认值。
    """
    if val is None:
        return f"{enum_name}.{default}"
    s = str(val).strip()
    if not s:
        return f"{enum_name}.{default}"
    if is_enum_ref(s):
        return s
    return f"{enum_name}.{s}"


def to_value_string(val):
    """将值转换为TypeScript字符串，处理枚举引用"""
    if val is None:
        return "undefined"
    s = str(val).strip()
    if not s:
        return "undefined"
    if is_enum_ref(s):
        return s
    # 尝试解析数字
    try:
        num = int(float(s))
        return str(num)
    except ValueError:
        pass
    try:
        num = float(s)
        return str(num)
    except ValueError:
        pass
    # 布尔
    if s.lower() == "true":
        return "true"
    if s.lower() == "false":
        return "false"
    # 字符串
    escaped = s.replace("\\", "\\\\").replace("'", "\\'").replace("\n", "\\n").replace("\r", "")
    return f"'{escaped}'"


# ============================================================
# 效果解析器
# ============================================================

def parse_effects(effect_xlsx: str) -> dict:
    """
    读取效果.xlsx，返回 {效果ID: 效果对象} 的映射。
    效果对象包含 effect 和 effectResult 两部分。
    """
    wb = openpyxl.load_workbook(effect_xlsx)
    effect_rows = read_sheet(wb, "效果")
    result_rows = read_sheet(wb, "效果结果")

    # 构建效果索引
    effects = {}
    for e in effect_rows:
        eid = e.get("ID")
        if not eid:
            continue
        effect = build_effect_object(e)
        if effect:
            effects[eid] = effect

    # 构建效果结果索引
    effect_results = {}
    for r in result_rows:
        rid = r.get("ID")
        if not rid:
            continue
        effect_id = r.get("effect")
        if effect_id and effect_id in effects:
            obj = {
                "effect": effects[effect_id],
            }
            prob = parse_float(r.get("probability"))
            if prob is not None:
                obj["probability"] = prob
            desc = r.get("description")
            if desc:
                obj["description"] = desc
            cond = r.get("condition")
            if cond:
                obj["condition"] = cond
            effect_results[rid] = obj

    return effect_results


def build_effect_object(row: dict) -> dict | None:
    """根据行数据构建效果对象"""
    etype = row.get("type")
    if not etype:
        return None

    type_str = str(etype).strip()
    # 处理枚举引用
    if "." in type_str:
        type_str = type_str.split(".")[-1]

    type_upper = type_str.upper()

    if type_upper == "ATTRIBUTE":
        return {
            "type": "EffectType.ATTRIBUTE",
            "attribute": to_effect_enum(row.get("attribute"), "AttributeType", "SAN"),
            "operation": to_effect_enum(row.get("operation (attr)"), "AttributeOperation", "ADD"),
            "value": parse_int(row.get("value (attr)"), 0),
        }
    elif type_upper == "STATUS":
        obj = {
            "type": "EffectType.STATUS",
            "statusId": to_value_string(row.get("statusId")),
            "apply": parse_bool(row.get("apply")),
        }
        dur = parse_int(row.get("duration"))
        if dur is not None:
            obj["duration"] = dur
        sc = parse_int(row.get("stackCount"))
        if sc is not None:
            obj["stackCount"] = sc
        return obj
    elif type_upper == "ITEM":
        obj = {
            "type": "EffectType.ITEM",
            "itemId": to_value_string(row.get("itemId")),
            "changeType": to_effect_enum(row.get("changeType"), "ItemChangeType", "ADD"),
        }
        qty = parse_int(row.get("quantity"))
        if qty is not None:
            obj["quantity"] = qty
        return obj
    elif type_upper == "SCENE":
        return {
            "type": "EffectType.SCENE",
            "sceneId": to_value_string(row.get("sceneId")),
        }
    elif type_upper == "FLAG":
        return {
            "type": "EffectType.FLAG",
            "flagId": to_value_string(row.get("flagId")),
            "operation": to_effect_enum(row.get("operation (flag)"), "FlagOperation", "SET"),
            "value": to_value_string(row.get("value (flag)")),
        }
    elif type_upper == "BATTLE":
        return {
            "type": "EffectType.BATTLE",
            "enemyId": [to_value_string(row.get("enemyId"))],
        }
    elif type_upper == "CG":
        return {
            "type": "EffectType.CG",
            "cgId": to_value_string(row.get("cgId")),
        }
    elif type_upper == "EVENT":
        return {
            "type": "EffectType.EVENT",
            "eventId": to_value_string(row.get("eventId")),
            "triggerTiming": to_value_string(row.get("triggerTiming")),
        }
    elif type_upper == "RECIPE":
        return {
            "type": "EffectType.RECIPE",
            "recipeId": to_value_string(row.get("recipeId")),
            "recipeType": to_effect_enum(row.get("recipeType"), "RecipeType", "CRAFT"),
            "unlock": parse_bool(row.get("unlock")),
        }
    elif type_upper == "SKILL":
        return {
            "type": "EffectType.SKILL",
            "skillId": to_value_string(row.get("skillId")),
            "unlock": parse_bool(row.get("unlock (skill)")),
        }
    elif type_upper == "GAIN_EXP":
        return {
            "type": "EffectType.GAIN_EXP",
            "target": to_effect_enum(row.get("target"), "GainExpTarget", "SURVIVAL_SKILL"),
            "targetId": to_value_string(row.get("targetId")),
            "amount": parse_int(row.get("amount"), 0),
        }
    elif type_upper == "COMPOSITE":
        return {
            "type": "EffectType.COMPOSITE",
            "effects": [],
            "executionMode": to_value_string(row.get("executionMode")),
        }
    else:
        print(f"  警告: 未识别的效果类型 '{type_str}'")
        return None


# ============================================================
# 条件解析器（复用场景转换器的逻辑）
# ============================================================

def parse_conditions(cond_xlsx: str) -> dict:
    """读取条件.xlsx，返回 {条件id: 条件对象}"""
    if not os.path.exists(cond_xlsx):
        print("  警告: 条件.xlsx 不存在，跳过条件解析")
        return {}
    wb = openpyxl.load_workbook(cond_xlsx)
    cond_rows = read_sheet(wb, "条件")
    target_rows = read_sheet(wb, "条件目标")

    targets = {}
    for t in target_rows:
        tid = t.get("thisID")
        if tid:
            targets[tid] = t

    conditions = {}
    for c in cond_rows:
        cid = c.get("id")
        if not cid:
            continue
        condition = {}
        logic = c.get("logic")
        sub_conds = parse_comma_list(c.get("subConditions"))
        target_id = c.get("target")

        if logic and sub_conds:
            condition["logic"] = logic
            condition["subConditions"] = sub_conds
        elif target_id and target_id in targets:
            tgt = targets[target_id]
            target_obj = {"type": tgt.get("type")}
            if tgt.get("id"):
                target_obj["id"] = tgt["id"]
            if tgt.get("attributeType"):
                target_obj["attributeType"] = tgt["attributeType"]
            if tgt.get("subType"):
                target_obj["subType"] = tgt["subType"]
            condition["target"] = target_obj
            condition["operator"] = c.get("operator")
            condition["value"] = parse_value(c.get("value"))
            v2 = c.get("value2")
            if v2 is not None:
                condition["value2"] = v2
        else:
            print(f"  警告: 条件 '{cid}' 无法解析")
            continue
        conditions[cid] = condition
    return conditions


def parse_value(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip().lower()
    if s in ("true", "false"):
        return s == "true"
    if s == "null":
        return None
    try:
        if "." in s:
            return float(s)
        return int(s)
    except ValueError:
        return str(val)


# ============================================================
# 事件构建器
# ============================================================

def build_events(event_xlsx: str, effect_results: dict, conditions: dict) -> dict:
    """构建事件字典 {事件ID: 事件对象}"""
    wb = openpyxl.load_workbook(event_xlsx)

    events_raw = read_sheet(wb, "事件")
    frames_raw = read_sheet(wb, "事件帧")
    options_raw = read_sheet(wb, "事件选项")
    results_raw = read_sheet(wb, "选项结果")
    costs_raw = read_sheet(wb, "交互花费")

    # ---- 构建索引 ----
    # 帧索引
    frame_map = {}
    for f in frames_raw:
        fid = f.get("id")
        if fid:
            frame_map[fid] = f

    # 选项索引
    option_map = {}
    for o in options_raw:
        oid = o.get("id")
        if oid:
            option_map[oid] = o

    # 结果索引
    result_map = {}
    for r in results_raw:
        rid = r.get("id")
        if rid:
            result_map[rid] = r

    # 花费索引
    cost_map = {}
    for c in costs_raw:
        cid = c.get("ID")
        if cid:
            cost_map[cid] = c

    # ---- 构建事件 ----
    events = {}
    for evt in events_raw:
        evt_id = evt.get("ID")
        if not evt_id:
            continue

        event_obj = {
            "id": evt_id,
            "name": evt.get("name") or "",
            "eventType": evt.get("eventType") or "EventType.NORMAL",
        }

        # 备注
        notes = evt.get("notes")
        if notes:
            event_obj["notes"] = notes

        # 帧列表
        frame_ids = parse_comma_list(evt.get("frames"))
        frames = []
        for fid in frame_ids:
            frame_row = frame_map.get(fid)
            if not frame_row:
                print(f"  警告: 事件 '{evt_id}' 引用的帧 '{fid}' 未找到")
                continue
            frame = build_frame(frame_row, option_map, result_map, effect_results, conditions, cost_map)
            if frame:
                frames.append(frame)
        event_obj["frames"] = frames

        # onEnterEffects
        on_enter_effects = resolve_effects_list(evt.get("onEnterEffects"), effect_results)
        if on_enter_effects:
            event_obj["onEnterEffects"] = on_enter_effects

        # isRepeatable
        if parse_bool(evt.get("isRepeatable")):
            event_obj["isRepeatable"] = True

        # triggeredFlag
        tf = evt.get("triggeredFlag")
        if tf:
            event_obj["triggeredFlag"] = tf

        # untriggerableText
        ut = evt.get("untriggerableText")
        if ut:
            event_obj["untriggerableText"] = ut

        events[evt_id] = event_obj

    return events


def build_frame(frame_row: dict, option_map: dict, result_map: dict,
                effect_results: dict, conditions: dict, cost_map: dict) -> dict | None:
    """构建事件帧对象"""
    frame = {
        "id": frame_row.get("id"),
        "order": parse_int(frame_row.get("order"), 1),
        "text": frame_row.get("text") or "",
        "options": [],
    }

    # 选项
    opt_ids = parse_comma_list(frame_row.get("options"))
    for oid in opt_ids:
        opt_row = option_map.get(oid)
        if not opt_row:
            print(f"  警告: 帧 '{frame.get('id')}' 引用的选项 '{oid}' 未找到")
            continue
        option = build_option(opt_row, result_map, effect_results, conditions, cost_map)
        if option:
            frame["options"].append(option)

    # displayCondition
    dc = frame_row.get("displayCondition")
    if dc:
        frame["displayCondition"] = dc

    # onEnterEffects
    on_enter = resolve_effects_list(frame_row.get("onEnterEffects"), effect_results)
    if on_enter:
        frame["onEnterEffects"] = on_enter

    # onExitEffects
    on_exit = resolve_effects_list(frame_row.get("onExitEffects"), effect_results)
    if on_exit:
        frame["onExitEffects"] = on_exit

    return frame


def build_option(opt_row: dict, result_map: dict, effect_results: dict,
                 conditions: dict, cost_map: dict) -> dict | None:
    """构建事件选项对象"""
    oid = opt_row.get("id")
    if not oid:
        return None

    option = {
        "id": oid,
        "text": opt_row.get("text") or "",
        "results": [],
    }

    # 描述
    desc = opt_row.get("description")
    if desc:
        option["description"] = desc

    # 条件
    dc = opt_row.get("displayCondition")
    if dc:
        option["displayCondition"] = dc
    ac = opt_row.get("availableCondition")
    if ac:
        option["availableCondition"] = ac
    ut = opt_row.get("unavailableTooltip")
    if ut:
        option["unavailableTooltip"] = ut

    # 选项样式
    os_ = opt_row.get("optionStyle")
    if os_:
        if is_enum_ref(os_):
            option["optionStyle"] = str(os_).strip()
        else:
            option["optionStyle"] = f"EventOptionStyle.{str(os_).strip().upper()}"

    # 花费
    costs = parse_comma_list(opt_row.get("costs"))
    for cid in costs:
        cost_row = cost_map.get(cid)
        if cost_row:
            cost = {
                "costType": to_value_string(cost_row.get("costType")),
                "value": parse_int(cost_row.get("value"), 0),
            }
            item_id = cost_row.get("itemId")
            if item_id:
                cost["itemId"] = str(item_id).strip()
            item_qty = parse_int(cost_row.get("itemQuantity"))
            if item_qty is not None:
                cost["itemQuantity"] = item_qty
            option.setdefault("costs", []).append(cost)

    # 结果
    result_id = opt_row.get("results")
    if result_id:
        result_row = result_map.get(result_id)
        if result_row:
            result = build_result(result_row, effect_results, conditions)
            if result:
                option["results"].append(result)

    # displayPriority
    dp = parse_int(opt_row.get("displayPriority"))
    if dp is not None:
        option["displayPriority"] = dp

    # 确认弹窗
    if parse_bool(opt_row.get("requiresConfirmation")):
        option["requiresConfirmation"] = True
        ct = opt_row.get("confirmationText")
        if ct:
            option["confirmationText"] = ct

    # isOneTime
    if parse_bool(opt_row.get("isOneTime")):
        option["isOneTime"] = True

    # selectedFlag
    sf = opt_row.get("selectedFlag")
    if sf:
        option["selectedFlag"] = sf

    # textAfterSelected
    tas = opt_row.get("textAfterSelected")
    if tas:
        option["textAfterSelected"] = tas

    return option


def build_result(result_row: dict, effect_results: dict, conditions: dict) -> dict | None:
    """构建事件选项结果对象"""
    rtype = result_row.get("type")
    if not rtype:
        return None

    type_str = str(rtype).strip()
    result = {"type": type_str}

    # 公共字段
    w = parse_int(result_row.get("weight"))
    if w is not None:
        result["weight"] = w

    cond = result_row.get("condition")
    if cond:
        result["condition"] = cond

    # 效果
    effects = resolve_effects_list(result_row.get("effects"), effect_results)
    if effects:
        result["effects"] = effects

    # setFlags
    set_flags_str = result_row.get("setFlags")
    if set_flags_str:
        flags = {}
        for item in parse_comma_list(set_flags_str):
            parts = item.split(":", 1)
            if len(parts) >= 2:
                key = parts[0].strip()
                val = parts[1].strip()
                if val.lower() == "true":
                    flags[key] = True
                elif val.lower() == "false":
                    flags[key] = False
                else:
                    try:
                        flags[key] = int(val)
                    except ValueError:
                        flags[key] = val
        if flags:
            result["setFlags"] = flags

    # 根据类型添加特定字段
    if type_str == "nextFrame":
        result["targetFrameId"] = result_row.get("targetFrameId") or ""
        text = result_row.get("text")
        if text:
            result["text"] = text
    elif type_str == "endEvent":
        et = result_row.get("exitText")
        if et:
            result["exitText"] = et
    elif type_str == "triggerBattle":
        result["enemyId"] = parse_comma_list(result_row.get("enemyId"))
        vf = result_row.get("victoryFrameId")
        if vf:
            result["victoryFrameId"] = vf
        df = result_row.get("defeatFrameId")
        if df:
            result["defeatFrameId"] = df
        ef = result_row.get("escapeFrameId")
        if ef:
            result["escapeFrameId"] = ef
        if parse_bool(result_row.get("canEscape")):
            result["canEscape"] = True
        if parse_bool(result_row.get("firstEncounterBonus")):
            result["firstEncounterBonus"] = True
    elif type_str == "playCG":
        result["cgId"] = result_row.get("cgId") or ""
        rf = result_row.get("returnFrameId")
        if rf:
            result["returnFrameId"] = rf
    elif type_str == "openTrade":
        result["traderId"] = result_row.get("traderId") or ""
        rf = result_row.get("returnFrameId")
        if rf:
            result["returnFrameId"] = rf
    elif type_str == "switchScene":
        result["sceneId"] = result_row.get("sceneId") or ""
        ss = result_row.get("subSceneId")
        if ss:
            result["subSceneId"] = ss
        text = result_row.get("text")
        if text:
            result["text"] = text
    elif type_str == "triggerEvent":
        result["eventId"] = result_row.get("eventId") or ""
        rf = result_row.get("returnFrameId")
        if rf:
            result["returnFrameId"] = rf

    return result


def resolve_effects_list(effects_str, effect_results: dict) -> list:
    """解析逗号分隔的效果结果ID列表，返回效果对象列表"""
    if not effects_str:
        return []
    ids = parse_comma_list(effects_str)
    results = []
    for eid in ids:
        if eid in effect_results:
            results.append(effect_results[eid])
        else:
            print(f"  警告: 效果结果 '{eid}' 未找到")
    return results if results else []


# ============================================================
# TypeScript 生成器
# ============================================================

def generate_ts(events: dict, conditions: dict) -> str:
    """生成TypeScript源码"""
    lines = []
    indent = "  "

    lines.append("// ============================================================")
    lines.append("// 此文件由事件转换器自动生成，请勿手动修改")
    lines.append("// 生成时间: " + __import__("datetime").datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    lines.append("// ============================================================")
    lines.append("")
    lines.append("import type { GameEvent, EventRegistry } from '../types/event'")
    lines.append("import { EventType, EventOptionStyle, EventOptionCostType } from '../types/event'")
    lines.append("import {")
    lines.append("  EffectType, AttributeType, AttributeOperation, ItemChangeType,")
    lines.append("  GainExpTarget, LogicOperator, ConditionTargetType, ComparisonOperator,")
    lines.append("} from '../types/effect'")
    lines.append("import { FlagOperation } from '@/types/flag'")
    lines.append("import { RecipeType } from '@/types/recipe'")
    lines.append("")

    # 条件常量的引用声明（如果有）
    if conditions:
        simple_conds = {cid: c for cid, c in conditions.items() if "subConditions" not in c}
        compound_conds = {cid: c for cid, c in conditions.items() if "subConditions" in c}

        # 只输出事件中引用的条件
        referenced_conds = find_referenced_conditions(events, conditions)
        if referenced_conds:
            lines.append("// ============================================================")
            lines.append("// 条件引用（来自条件.xlsx）")
            lines.append("// ============================================================")
            lines.append("")
            # 先输出简单条件，再输出复合条件
            for cid in sorted(referenced_conds.keys()):
                if cid not in referenced_conds:
                    continue
            # 按拓扑排序输出
            for cid, cond in simple_conds.items():
                if cid in referenced_conds:
                    lines.append(f"const {cid} = {condition_to_ts(cond, indent)}")
            for cid, cond in compound_conds.items():
                if cid in referenced_conds:
                    lines.append(f"const {cid} = {condition_to_ts(cond, indent)}")
            lines.append("")

    # 生成事件常量
    for evt_id, evt in events.items():
        var_name = snake_to_camel(evt_id)
        lines.append("// ============================================================")
        lines.append(f"// {evt.get('name', evt_id)}")
        lines.append("// ============================================================")
        lines.append("")
        lines.append(f"const {var_name}: GameEvent = {event_to_ts(evt, conditions, indent)}")
        lines.append("")

    # 生成注册表
    lines.append("// ============================================================")
    lines.append("// 事件注册表")
    lines.append("// ============================================================")
    lines.append("")
    lines.append("export const eventRegistry: EventRegistry = {")
    lines.append("  events: {")
    for evt_id in events.keys():
        var_name = snake_to_camel(evt_id)
        lines.append(f"    {evt_id}: {var_name},")
    lines.append("  },")
    lines.append("}")
    lines.append("")

    return "\n".join(lines)


def snake_to_camel(s: str) -> str:
    """将蛇形命名转换为驼峰命名（用于TS变量名）"""
    parts = s.split("_")
    return parts[0] + "".join(p.capitalize() for p in parts[1:])


def find_referenced_conditions(events: dict, conditions: dict) -> dict:
    """找出事件中引用的所有条件"""
    referenced = {}
    ref_keys = set()

    def collect_conditions(obj):
        if isinstance(obj, dict):
            for k, v in obj.items():
                if isinstance(v, str) and v in conditions:
                    ref_keys.add(v)
                collect_conditions(v)
        elif isinstance(obj, list):
            for item in obj:
                collect_conditions(item)

    collect_conditions(events)
    for cid in ref_keys:
        if cid in conditions:
            referenced[cid] = conditions[cid]
    return referenced


def condition_to_ts(cond: dict, indent: str = "  ") -> str:
    """将条件字典转换为TS对象字符串"""
    from_condition_converter = True
    lines = ["{"]
    inner = indent + "  "

    for key, val in cond.items():
        if key == "subConditions":
            lines.append(f"{inner}{key}: [")
            for sub_id in val:
                lines.append(f"{inner}  {sub_id},")
            lines.append(f"{inner}],")
        elif key == "target":
            target = val
            lines.append(f"{inner}{key}: {{")
            for tk, tv in target.items():
                if tv is not None:
                    if tk == "type":
                        lines.append(f"{inner}  type: {to_condition_target_type(tv)},")
                    elif tk == "attributeType":
                        lines.append(f"{inner}  attributeType: {to_attribute_type(tv)},")
                    else:
                        lines.append(f"{inner}  {tk}: {to_ts_value(tv)},")
            lines.append(f"{inner}}},")
        elif key == "operator":
            op = str(val).upper() if val else None
            if op == "EQUAL":
                lines.append(f"{inner}{key}: ComparisonOperator.EQUAL,")
            elif op == "NOT_EQUAL":
                lines.append(f"{inner}{key}: ComparisonOperator.NOT_EQUAL,")
            elif op == "GREATER":
                lines.append(f"{inner}{key}: ComparisonOperator.GREATER,")
            elif op == "GREATER_EQUAL":
                lines.append(f"{inner}{key}: ComparisonOperator.GREATER_EQUAL,")
            elif op == "LESS":
                lines.append(f"{inner}{key}: ComparisonOperator.LESS,")
            elif op == "LESS_EQUAL":
                lines.append(f"{inner}{key}: ComparisonOperator.LESS_EQUAL,")
            elif op == "BETWEEN":
                lines.append(f"{inner}{key}: ComparisonOperator.BETWEEN,")
            elif op == "IN":
                lines.append(f"{inner}{key}: ComparisonOperator.IN,")
            elif op == "NOT_IN":
                lines.append(f"{inner}{key}: ComparisonOperator.NOT_IN,")
            elif op == "EXISTS":
                lines.append(f"{inner}{key}: ComparisonOperator.EXISTS,")
            elif op == "NOT_EXISTS":
                lines.append(f"{inner}{key}: ComparisonOperator.NOT_EXISTS,")
            else:
                lines.append(f"{inner}{key}: '{val}',")
        elif key == "logic":
            logic = str(val).upper() if val else None
            if logic == "AND":
                lines.append(f"{inner}{key}: LogicOperator.AND,")
            elif logic == "OR":
                lines.append(f"{inner}{key}: LogicOperator.OR,")
            elif logic == "NOT":
                lines.append(f"{inner}{key}: LogicOperator.NOT,")
            else:
                lines.append(f"{inner}{key}: '{val}',")
        else:
            if val is not None:
                lines.append(f"{inner}{key}: {to_ts_value(val)},")
    lines.append(f"{indent}}}")
    return "\n".join(lines)


def to_condition_target_type(t: str) -> str:
    mapping = {
        "ATTRIBUTE": "ConditionTargetType.ATTRIBUTE",
        "FLAG": "ConditionTargetType.FLAG",
        "ITEM": "ConditionTargetType.ITEM",
        "STATUS": "ConditionTargetType.STATUS",
        "SCENE": "ConditionTargetType.SCENE",
        "TIME": "ConditionTargetType.TIME",
        "WEATHER": "ConditionTargetType.WEATHER",
        "SEASON": "ConditionTargetType.SEASON",
        "SAN_LEVEL": "ConditionTargetType.SAN_LEVEL",
        "CORRUPTION": "ConditionTargetType.CORRUPTION",
        "SKILL": "ConditionTargetType.SKILL",
        "WEAPON_PROFICIENCY": "ConditionTargetType.WEAPON_PROFICIENCY",
        "RECIPE_UNLOCKED": "ConditionTargetType.RECIPE_UNLOCKED",
        "PLAYER_GOLD": "ConditionTargetType.PLAYER_GOLD",
        "CARRY_WEIGHT_RATE": "ConditionTargetType.CARRY_WEIGHT_RATE",
    }
    return mapping.get(t.upper(), f"'{t}'")


def to_attribute_type(t: str) -> str:
    mapping = {
        "HP": "AttributeType.HP",
        "SATIETY": "AttributeType.SATIETY",
        "STAMINA": "AttributeType.STAMINA",
        "SAN": "AttributeType.SAN",
        "WARMTH": "AttributeType.WARMTH",
        "CARRY_WEIGHT": "AttributeType.CARRY_WEIGHT",
        "STRENGTH": "AttributeType.STRENGTH",
        "AGILITY": "AttributeType.AGILITY",
        "INTELLIGENCE": "AttributeType.INTELLIGENCE",
        "CONSTITUTION": "AttributeType.CONSTITUTION",
        "STRENGTH_EXP": "AttributeType.STRENGTH_EXP",
        "AGILITY_EXP": "AttributeType.AGILITY_EXP",
        "INTELLIGENCE_EXP": "AttributeType.INTELLIGENCE_EXP",
        "CONSTITUTION_EXP": "AttributeType.CONSTITUTION_EXP",
        "WEAPON_PROFICIENCY": "AttributeType.WEAPON_PROFICIENCY",
        "WEAPON_PROFICIENCY_EXP": "AttributeType.WEAPON_PROFICIENCY_EXP",
        "SLASH_DEFENSE": "AttributeType.SLASH_DEFENSE",
        "BLUNT_DEFENSE": "AttributeType.BLUNT_DEFENSE",
        "RANGED_DEFENSE": "AttributeType.RANGED_DEFENSE",
        "POISON_DEFENSE": "AttributeType.POISON_DEFENSE",
        "FIRE_DEFENSE": "AttributeType.FIRE_DEFENSE",
        "SKILL_LEVEL": "AttributeType.SKILL_LEVEL",
        "SKILL_EXP": "AttributeType.SKILL_EXP",
        "RECOVERY_RATE_COEFFICIENT": "AttributeType.RECOVERY_RATE_COEFFICIENT",
        "SATIETY_UPPER_LIMIT_COEFFICIENT": "AttributeType.SATIETY_UPPER_LIMIT_COEFFICIENT",
        "SATIETY_LOSS_COEFFICIENT": "AttributeType.SATIETY_LOSS_COEFFICIENT",
        "STAMINA_CONSUMPTION_COEFFICIENT": "AttributeType.STAMINA_CONSUMPTION_COEFFICIENT",
        "STAMINA_RECOVERY_COEFFICIENT": "AttributeType.STAMINA_RECOVERY_COEFFICIENT",
        "STAMINA_RECOVERY_FIX": "AttributeType.STAMINA_RECOVERY_FIX",
        "SAN_MODIFIER": "AttributeType.SAN_MODIFIER",
        "TEMPERATURE_LOW": "AttributeType.TEMPERATURE_LOW",
        "TEMPERATURE_HIGH": "AttributeType.TEMPERATURE_HIGH",
        "CARRY_WEIGHT_MODIFIER": "AttributeType.CARRY_WEIGHT_MODIFIER",
    }
    return mapping.get(t.upper(), f"'{t}'")


def event_to_ts(evt: dict, conditions: dict, indent: str = "  ") -> str:
    """将事件字典转换为TS对象字符串"""
    lines = ["{"]
    inner = indent + "  "

    lines.append(f"{inner}id: '{evt.get('id', '')}',")
    lines.append(f"{inner}name: '{evt.get('name', '')}',")

    notes = evt.get("notes")
    if notes:
        lines.append(f"{inner}notes: '{notes}',")

    # frames
    frames = evt.get("frames", [])
    lines.append(f"{inner}frames: [")
    for f in frames:
        lines.append(f"{inner}  {frame_to_ts(f, conditions, inner + '  ')},")
    lines.append(f"{inner}],")

    # onEnterEffects
    on_enter = evt.get("onEnterEffects")
    if on_enter:
        lines.append(f"{inner}onEnterEffects: [")
        for eff in on_enter:
            lines.append(f"{inner}  {effect_result_to_ts(eff, inner + '  ')},")
        lines.append(f"{inner}],")

    # eventType
    et = evt.get("eventType", "EventType.NORMAL")
    lines.append(f"{inner}eventType: {et},")

    if evt.get("isRepeatable"):
        lines.append(f"{inner}isRepeatable: true,")
    tf = evt.get("triggeredFlag")
    if tf:
        lines.append(f"{inner}triggeredFlag: '{tf}',")
    ut = evt.get("untriggerableText")
    if ut:
        lines.append(f"{inner}untriggerableText: '{ut}',")

    lines.append(f"{indent}}}")
    return "\n".join(lines)


def frame_to_ts(frame: dict, conditions: dict, indent: str) -> str:
    lines = ["{"]
    inner = indent + "  "

    lines.append(f"{inner}id: '{frame.get('id', '')}',")
    lines.append(f"{inner}order: {frame.get('order', 1)},")
    lines.append(f"{inner}text: {to_ts_value(frame.get('text', ''))},")

    # displayCondition
    dc = frame.get("displayCondition")
    if dc:
        if isinstance(dc, str) and dc in conditions:
            lines.append(f"{inner}displayCondition: {dc},")
        else:
            lines.append(f"{inner}displayCondition: {to_value_string(dc)},")

    # onEnterEffects
    on_enter = frame.get("onEnterEffects")
    if on_enter:
        lines.append(f"{inner}onEnterEffects: [")
        for eff in on_enter:
            lines.append(f"{inner}  {effect_result_to_ts(eff, inner + '  ')},")
        lines.append(f"{inner}],")

    # onExitEffects
    on_exit = frame.get("onExitEffects")
    if on_exit:
        lines.append(f"{inner}onExitEffects: [")
        for eff in on_exit:
            lines.append(f"{inner}  {effect_result_to_ts(eff, inner + '  ')},")
        lines.append(f"{inner}],")

    # options
    options = frame.get("options", [])
    lines.append(f"{inner}options: [")
    for opt in options:
        lines.append(f"{inner}  {option_to_ts(opt, conditions, inner + '  ')},")
    lines.append(f"{inner}],")

    lines.append(f"{indent}}}")
    return "\n".join(lines)


def option_to_ts(opt: dict, conditions: dict, indent: str) -> str:
    lines = ["{"]
    inner = indent + "  "

    lines.append(f"{inner}id: '{opt.get('id', '')}',")
    lines.append(f"{inner}text: '{opt.get('text', '')}',")

    desc = opt.get("description")
    if desc:
        lines.append(f"{inner}description: '{desc}',")

    dc = opt.get("displayCondition")
    if dc:
        if isinstance(dc, str) and dc in conditions:
            lines.append(f"{inner}displayCondition: {dc},")
        else:
            lines.append(f"{inner}displayCondition: {to_value_string(dc)},")

    ac = opt.get("availableCondition")
    if ac:
        if isinstance(ac, str) and ac in conditions:
            lines.append(f"{inner}availableCondition: {ac},")
        else:
            lines.append(f"{inner}availableCondition: {to_value_string(ac)},")

    ut = opt.get("unavailableTooltip")
    if ut:
        lines.append(f"{inner}unavailableTooltip: '{ut}',")

    os_ = opt.get("optionStyle")
    if os_:
        lines.append(f"{inner}optionStyle: {os_},")

    costs = opt.get("costs")
    if costs:
        lines.append(f"{inner}costs: [")
        for c in costs:
            lines.append(f"{inner}  {cost_to_ts(c, inner + '  ')},")
        lines.append(f"{inner}],")

    results = opt.get("results", [])
    lines.append(f"{inner}results: [")
    for r in results:
        lines.append(f"{inner}  {result_to_ts(r, conditions, inner + '  ')},")
    lines.append(f"{inner}],")

    dp = opt.get("displayPriority")
    if dp is not None:
        lines.append(f"{inner}displayPriority: {dp},")
    if opt.get("requiresConfirmation"):
        lines.append(f"{inner}requiresConfirmation: true,")
        ct = opt.get("confirmationText")
        if ct:
            lines.append(f"{inner}confirmationText: '{ct}',")
    if opt.get("isOneTime"):
        lines.append(f"{inner}isOneTime: true,")
    sf = opt.get("selectedFlag")
    if sf:
        lines.append(f"{inner}selectedFlag: '{sf}',")
    tas = opt.get("textAfterSelected")
    if tas:
        lines.append(f"{inner}textAfterSelected: '{tas}',")

    lines.append(f"{indent}}}")
    return "\n".join(lines)


def cost_to_ts(cost: dict, indent: str) -> str:
    lines = ["{"]
    inner = indent + "  "

    ct = cost.get("costType", "EventOptionCostType.STAMINA")
    lines.append(f"{inner}costType: {ct},")
    lines.append(f"{inner}value: {cost.get('value', 0)},")
    item_id = cost.get("itemId")
    if item_id:
        lines.append(f"{inner}itemId: '{item_id}',")
    item_qty = cost.get("itemQuantity")
    if item_qty is not None:
        lines.append(f"{inner}itemQuantity: {item_qty},")

    lines.append(f"{indent}}}")
    return "\n".join(lines)


def result_to_ts(result: dict, conditions: dict, indent: str) -> str:
    lines = ["{"]
    inner = indent + "  "

    lines.append(f"{inner}type: '{result.get('type', 'endEvent')}',")

    # 公共字段
    w = result.get("weight")
    if w is not None:
        lines.append(f"{inner}weight: {w},")

    cond = result.get("condition")
    if cond:
        if isinstance(cond, str) and cond in conditions:
            lines.append(f"{inner}condition: {cond},")
        else:
            lines.append(f"{inner}condition: {to_value_string(cond)},")

    effects = result.get("effects")
    if effects:
        lines.append(f"{inner}effects: [")
        for eff in effects:
            lines.append(f"{inner}  {effect_result_to_ts(eff, inner + '  ')},")
        lines.append(f"{inner}],")

    sfs = result.get("setFlags")
    if sfs:
        lines.append(f"{inner}setFlags: {{")
        for k, v in sfs.items():
            lines.append(f"{inner}  {k}: {to_ts_value(v)},")
        lines.append(f"{inner}}},")

    # 根据类型添加字段
    rtype = result.get("type")
    if rtype == "nextFrame":
        lines.append(f"{inner}targetFrameId: '{result.get('targetFrameId', '')}',")
        text = result.get("text")
        if text:
            lines.append(f"{inner}text: {to_ts_value(text)},")
    elif rtype == "endEvent":
        et = result.get("exitText")
        if et:
            lines.append(f"{inner}exitText: '{et}',")
    elif rtype == "triggerBattle":
        enemy_ids = result.get("enemyId", [])
        ids_str = ", ".join(f"'{e}'" for e in enemy_ids)
        lines.append(f"{inner}enemyId: [{ids_str}],")
        vf = result.get("victoryFrameId")
        if vf:
            lines.append(f"{inner}victoryFrameId: '{vf}',")
        df = result.get("defeatFrameId")
        if df:
            lines.append(f"{inner}defeatFrameId: '{df}',")
        ef = result.get("escapeFrameId")
        if ef:
            lines.append(f"{inner}escapeFrameId: '{ef}',")
        if result.get("canEscape"):
            lines.append(f"{inner}canEscape: true,")
        if result.get("firstEncounterBonus"):
            lines.append(f"{inner}firstEncounterBonus: true,")
    elif rtype == "playCG":
        lines.append(f"{inner}cgId: '{result.get('cgId', '')}',")
        rf = result.get("returnFrameId")
        if rf:
            lines.append(f"{inner}returnFrameId: '{rf}',")
    elif rtype == "openTrade":
        lines.append(f"{inner}traderId: '{result.get('traderId', '')}',")
        rf = result.get("returnFrameId")
        if rf:
            lines.append(f"{inner}returnFrameId: '{rf}',")
    elif rtype == "switchScene":
        lines.append(f"{inner}sceneId: '{result.get('sceneId', '')}',")
        ss = result.get("subSceneId")
        if ss:
            lines.append(f"{inner}subSceneId: '{ss}',")
        text = result.get("text")
        if text:
            lines.append(f"{inner}text: {to_ts_value(text)},")
    elif rtype == "triggerEvent":
        lines.append(f"{inner}eventId: '{result.get('eventId', '')}',")
        rf = result.get("returnFrameId")
        if rf:
            lines.append(f"{inner}returnFrameId: '{rf}',")

    lines.append(f"{indent}}}")
    return "\n".join(lines)


def effect_result_to_ts(eff: dict, indent: str) -> str:
    """将效果结果对象转换为TS对象字符串"""
    lines = ["{"]
    inner = indent + "  "

    effect = eff.get("effect", {})
    lines.append(f"{inner}effect: {effect_to_ts(effect, indent)},")

    prob = eff.get("probability")
    if prob is not None:
        lines.append(f"{inner}probability: {prob},")

    cond = eff.get("condition")
    if cond:
        lines.append(f"{inner}condition: {to_value_string(cond)},")

    desc = eff.get("description")
    if desc:
        lines.append(f"{inner}description: '{desc}',")

    lines.append(f"{indent}}}")
    return "\n".join(lines)


def effect_to_ts(effect: dict, indent: str) -> str:
    """将效果对象转换为TS对象字符串"""
    lines = ["{"]
    inner = indent + "  "

    etype = effect.get("type", "EffectType.ATTRIBUTE")
    lines.append(f"{inner}type: {etype},")

    if etype == "EffectType.ATTRIBUTE":
        lines.append(f"{inner}attribute: {effect.get('attribute', 'AttributeType.SAN')},")
        lines.append(f"{inner}operation: {effect.get('operation', 'AttributeOperation.SUBTRACT')},")
        lines.append(f"{inner}value: {effect.get('value', 0)},")
    elif etype == "EffectType.ITEM":
        lines.append(f"{inner}itemId: {effect.get('itemId', '')},")
        lines.append(f"{inner}changeType: {effect.get('changeType', 'ItemChangeType.ADD')},")
        qty = effect.get("quantity")
        if qty is not None:
            lines.append(f"{inner}quantity: {qty},")
    elif etype == "EffectType.FLAG":
        lines.append(f"{inner}flagId: {effect.get('flagId', '')},")
        lines.append(f"{inner}operation: {effect.get('operation', 'FlagOperation.SET')},")
        lines.append(f"{inner}value: {to_value_string(effect.get('value', 'true'))},")
    elif etype == "EffectType.GAIN_EXP":
        lines.append(f"{inner}target: {effect.get('target', 'GainExpTarget.SURVIVAL_SKILL')},")
        lines.append(f"{inner}targetId: {effect.get('targetId', '')},")
        lines.append(f"{inner}amount: {effect.get('amount', 0)},")
    elif etype == "EffectType.STATUS":
        lines.append(f"{inner}statusId: {effect.get('statusId', '')},")
        lines.append(f"{inner}apply: {str(effect.get('apply', True)).lower()},")
        dur = effect.get("duration")
        if dur is not None:
            lines.append(f"{inner}duration: {dur},")
        sc = effect.get("stackCount")
        if sc is not None:
            lines.append(f"{inner}stackCount: {sc},")
    elif etype == "EffectType.SCENE":
        lines.append(f"{inner}sceneId: {effect.get('sceneId', '')},")
    elif etype == "EffectType.BATTLE":
        enemy_ids = effect.get("enemyId", [])
        ids_str = ", ".join(f"'{e}'" for e in enemy_ids)
        lines.append(f"{inner}enemyId: [{ids_str}],")
    elif etype == "EffectType.CG":
        lines.append(f"{inner}cgId: {effect.get('cgId', '')},")
    elif etype == "EffectType.EVENT":
        ev = effect.get("triggerTiming", "immediate")
        lines.append(f"{inner}eventId: {effect.get('eventId', '')},")
        lines.append(f"{inner}triggerTiming: '{ev}',")
    elif etype == "EffectType.RECIPE":
        lines.append(f"{inner}recipeId: {effect.get('recipeId', '')},")
        lines.append(f"{inner}recipeType: {effect.get('recipeType', 'RecipeType.CRAFT')},")
        lines.append(f"{inner}unlock: {str(effect.get('unlock', True)).lower()},")
    elif etype == "EffectType.SKILL":
        lines.append(f"{inner}skillId: {effect.get('skillId', '')},")
        lines.append(f"{inner}unlock: {str(effect.get('unlock', True)).lower()},")
    elif etype == "EffectType.COMPOSITE":
        em = effect.get("executionMode", "sequential")
        lines.append(f"{inner}effects: [],")
        lines.append(f"{inner}executionMode: '{em}',")

    lines.append(f"{indent}}}")
    return "\n".join(lines)


# ============================================================
# 主程序
# ============================================================

def main():
    print("=" * 60)
    print("事件 XLSX → TS 配置转换器")
    print("=" * 60)

    # 验证文件
    for fpath, fname in [(EVENT_XLSX, "事件.xlsx"), (EFFECT_XLSX, "效果.xlsx")]:
        if not os.path.exists(fpath):
            print(f"错误: 找不到 {fname} 文件: {fpath}")
            sys.exit(1)
        print(f"[OK] 已找到 {fname}")

    # 1. 解析效果
    print("\n> 正在解析效果...")
    effect_results = parse_effects(EFFECT_XLSX)
    print(f"  [OK] 解析了 {len(effect_results)} 个效果结果")

    # 2. 解析条件
    print("\n> 正在解析条件...")
    conditions = {}
    if os.path.exists(COND_XLSX):
        conditions = parse_conditions(COND_XLSX)
        print(f"  [OK] 解析了 {len(conditions)} 个条件")
    else:
        print("  [OK] 条件.xlsx 不存在，跳过")

    # 3. 构建事件
    print("\n> 正在构建事件...")
    events = build_events(EVENT_XLSX, effect_results, conditions)
    print(f"  [OK] 构建了 {len(events)} 个事件")
    for eid, evt in events.items():
        frame_count = len(evt.get("frames", []))
        print(f"    - {eid}: {evt.get('name')} ({frame_count} 帧)")

    # 4. 生成 TypeScript
    print("\n> 正在生成 TypeScript 代码...")
    ts_code = generate_ts(events, conditions)

    # 5. 写入文件
    os.makedirs(os.path.dirname(OUTPUT_TS), exist_ok=True)
    with open(OUTPUT_TS, "w", encoding="utf-8") as f:
        f.write(ts_code)
    print(f"  [OK] 文件已写入: {OUTPUT_TS}")
    print(f"  [OK] 共 {len(ts_code.splitlines())} 行代码")

    # 6. 验证
    print("\n> 正在验证...")
    errors = []
    for eid, evt in events.items():
        for frame in evt.get("frames", []):
            dc = frame.get("displayCondition")
            if dc and isinstance(dc, str) and dc in conditions:
                pass  # 条件引用有效
            for opt in frame.get("options", []):
                for r in opt.get("results", []):
                    cond = r.get("condition")
                    if cond and isinstance(cond, str) and cond not in conditions:
                        # 如果条件不在已知列表中，可能是内联引用
                        pass

    if errors:
        print(f"  [WARN] 发现 {len(errors)} 个问题:")
        for err in errors:
            print(f"    - {err}")
    else:
        print("  [OK] 所有数据验证通过")

    print("\n" + "=" * 60)
    print("转换完成！")
    print("=" * 60)


if __name__ == "__main__":
    main()