"""
场景XLSX转TS配置转换器
========================
从 场景.xlsx 和 条件.xlsx 读取数据，转换为符合 types/scene.ts 格式的 TypeScript 代码，
输出到 config/scenes.ts。

依赖：pip install openpyxl

使用方式：
    python "剧本设计/场景XLSX转TS配置.py"

数据关联规则：
  - 非列表类型关联：通过ID字段直接匹配（如条件->条件目标）
  - 列表类型关联：多个关联ID使用逗号分隔（如场景->描述列表、描述->事件入口）
  - 数据类型本身已包含ID字段时，直接使用该ID进行关联
"""

import os
import re
import sys
import json
from collections import defaultdict

# 控制台编码兼容（Windows GBK 环境）
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

try:
    import openpyxl
except ImportError:
    print("错误：需要 openpyxl 库。请运行: pip install openpyxl")
    sys.exit(1)


# ============================================================
# 路径配置
# ============================================================

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)  # RPGame-demo 目录

SCENE_XLSX = os.path.join(SCRIPT_DIR, "场景.xlsx")
COND_XLSX = os.path.join(SCRIPT_DIR, "条件.xlsx")
OUTPUT_TS = os.path.join(PROJECT_ROOT, "src", "config", "scenes.ts")


# ============================================================
# XLSX 读取工具
# ============================================================

def read_sheet(wb, sheet_name: str) -> list[dict]:
    """读取工作表，返回列表[行字典]，key为表头"""
    if sheet_name not in wb.sheetnames:
        print(f"  警告: 工作表 '{sheet_name}' 不存在，跳过")
        return []
    ws = wb[sheet_name]
    headers = []
    for c in range(1, ws.max_column + 1):
        raw = str(ws.cell(1, c).value or "").strip()
        # 取换行前的内容
        key = raw.split("\n")[0].strip()
        # 去除括号中的注释（如 "thisID(条件目标id)" -> "thisID"）
        key = re.sub(r"[（(][^）)]*[）)]", "", key).strip()
        headers.append(key)
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        has_data = False
        for c in range(1, ws.max_column + 1):
            val = ws.cell(r, c).value
            key = headers[c - 1]
            row[key] = val
            if val is not None:
                has_data = True
        if has_data:
            rows.append(row)
    return rows


def parse_comma_list(val) -> list[str]:
    """解析逗号分隔的列表，过滤空值和None"""
    if val is None:
        return []
    s = str(val).strip()
    if not s:
        return []
    return [item.strip() for item in s.split(",") if item.strip()]


def parse_bool(val) -> bool:
    """解析布尔值"""
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    return s in ("true", "yes", "1")


def parse_int(val, default=None):
    """解析整数"""
    if val is None:
        return default
    try:
        return int(float(str(val)))
    except (ValueError, TypeError):
        return default


def parse_float(val, default=None):
    """解析浮点数"""
    if val is None:
        return default
    try:
        return float(str(val))
    except (ValueError, TypeError):
        return default


def to_ts_value(val, indent=0):
    """将Python值转换为TypeScript值字符串"""
    if val is None:
        return "undefined"
    if isinstance(val, bool):
        return "true" if val else "false"
    if isinstance(val, int):
        return str(val)
    if isinstance(val, float):
        return str(val)
    if isinstance(val, str):
        # 数字字符串还是数字
        # 如果是json中的特殊关键字
        if val.lower() in ("true", "false"):
            return val.lower()
        if val.lower() == "null":
            return "undefined"
        # 尝试解析为数字
        try:
            if "." in val:
                return str(float(val))
            return str(int(val))
        except ValueError:
            pass
        # 转义字符串中的特殊字符
        escaped = (val.replace("\\", "\\\\")
                     .replace("'", "\\'")
                     .replace("\n", "\\n")
                     .replace("\r", ""))
        return f"'{escaped}'"
    # 列表或字典当作字面量
    return json.dumps(val, ensure_ascii=False)


# ============================================================
# 条件解析器
# ============================================================

def parse_conditions(cond_xlsx: str) -> dict:
    """
    读取条件.xlsx，返回 {条件id: 条件字典} 的映射。
    条件字典结构符合 types/effect.ts 中的 Condition 接口。
    """
    wb = openpyxl.load_workbook(cond_xlsx)
    cond_rows = read_sheet(wb, "条件")
    target_rows = read_sheet(wb, "条件目标")

    # 构建条件目标索引
    targets = {}
    for t in target_rows:
        tid = t.get("thisID")
        if tid:
            targets[tid] = t

    # 构建条件映射
    conditions = {}
    for c in cond_rows:
        cond_id = c.get("id")
        if not cond_id:
            continue

        condition = {}
        logic = c.get("logic")
        sub_conds = parse_comma_list(c.get("subConditions"))
        target_id = c.get("target")

        if logic and sub_conds:
            # 复合条件：AND / OR / NOT
            condition["logic"] = logic
            condition["subConditions"] = sub_conds
        elif target_id and target_id in targets:
            # 简单条件：target + operator + value
            tgt = targets[target_id]
            target_obj = {"type": tgt.get("type")}
            if tgt.get("id"):
                target_obj["id"] = tgt["id"]
            if tgt.get("attributeType"):
                target_obj["attributeType"] = tgt["attributeType"]
            if tgt.get("subType"):
                target_obj["subType"] = tgt["subType"]
            condition["target"] = target_obj
            condition["operator"] = c.get("operator")
            condition["value"] = parse_value(c.get("value"))
            v2 = c.get("value2")
            if v2 is not None:
                condition["value2"] = v2
        else:
            print(f"  警告: 条件 '{cond_id}' 无法解析（target '{target_id}' 未找到或逻辑不完整）")
            continue

        conditions[cond_id] = condition
    return conditions


def parse_value(val):
    """智能解析值：优先转为数字，保持字符串"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip().lower()
    if s in ("true", "false"):
        return s == "true"
    if s == "null":
        return None
    try:
        if "." in s:
            return float(s)
        return int(s)
    except ValueError:
        return str(val)


# ============================================================
# 场景数据解析器
# ============================================================

def build_scene_config(scene_xlsx: str, conditions: dict) -> dict:
    """
    读取场景.xlsx，构建完整的场景配置字典。
    返回结构符合 SceneRegistry 接口。
    """
    wb = openpyxl.load_workbook(scene_xlsx)

    # 读取所有工作表
    scenes_raw = read_sheet(wb, "场景")
    descs_raw = read_sheet(wb, "场景描述或事件入口")
    entries_raw = read_sheet(wb, "事件入口配置")
    interactions_raw = read_sheet(wb, "场景交互按钮")
    costs_raw = read_sheet(wb, "交互花费")
    bp_raw = read_sheet(wb, "交互行为参数")

    # ---- 构建索引 ----
    # 描述索引: desc_id -> row
    desc_map = {}
    for d in descs_raw:
        did = d.get("id")
        if did:
            desc_map[did] = d

    # 入口索引: 按 (parent_desc_id, key) 组织
    entry_map = defaultdict(list)
    for e in entries_raw:
        pid = e.get("ID")
        key = e.get("key")
        if pid and key:
            entry_map[pid].append(e)

    # 花费索引: cost_id -> row
    cost_map = {}
    for c in costs_raw:
        cid = c.get("ID")
        if cid:
            cost_map[cid] = c

    # 行为参数索引: bp_id -> row
    bp_map = {}
    for b in bp_raw:
        bid = b.get("ID")
        if bid:
            bp_map[bid] = b

    # 交互索引: 按 parent_scene_id 组织
    interaction_map = defaultdict(list)
    for i in interactions_raw:
        pid = i.get("ID")
        if pid:
            interaction_map[pid].append(i)

    # ---- 构建场景 ----
    scenes = {}
    sub_scenes = {}

    for s in scenes_raw:
        scene_id = s.get("id")
        if not scene_id:
            continue

        is_sub = parse_bool(s.get("是否子场景"))

        # 构建描述列表
        desc_ids = parse_comma_list(s.get("descriptions"))
        descriptions = []
        for did in desc_ids:
            desc_row = desc_map.get(did)
            if not desc_row:
                print(f"  警告: 场景 '{scene_id}' 引用的描述 '{did}' 未找到")
                continue

            desc = {
                "id": did,
                "priority": parse_int(desc_row.get("priority"), 0),
                "text": desc_row.get("text") or "",
                "isAutoTrigger": parse_bool(desc_row.get("isAutoTrigger")),
                "isOneTime": parse_bool(desc_row.get("isOneTime")),
            }

            # 权重（同优先级时按权重比例选择）
            weight = parse_int(desc_row.get("weight"))
            if weight is not None:
                desc["weight"] = weight

            # 显示条件
            cond_id = desc_row.get("displayCondition")
            if cond_id and cond_id in conditions:
                desc["displayCondition"] = cond_id
            elif cond_id and cond_id not in conditions:
                # 可能已内联，直接引用条件ID
                desc["displayCondition"] = cond_id

            # 事件入口
            event_entry_keys = parse_comma_list(desc_row.get("eventEntries"))
            entries = []
            for key in event_entry_keys:
                matching_entries = [e for e in entry_map.get(did, []) if e.get("key") == key]
                # 如果没找到，尝试用key直接在entry_map中查找
                if not matching_entries:
                    matching_entries = [e for e in entry_map.get(did, [])]
                    # 如果唯一匹配，就用它
                    if len(matching_entries) == 1:
                        pass
                    else:
                        matching_entries = []
                for e_row in matching_entries:
                    entry = {
                        "key": key,
                        "displayText": e_row.get("displayText") or "",
                        "eventId": e_row.get("eventId") or "",
                    }
                    # 可选字段
                    dcond = e_row.get("displayCondition")
                    if dcond:
                        entry["displayCondition"] = dcond
                    acond = e_row.get("availableCondition")
                    if acond:
                        entry["availableCondition"] = acond
                    if parse_bool(e_row.get("removeAfterClick")):
                        entry["removeAfterClick"] = True
                    cf = e_row.get("clickFlag")
                    if cf:
                        entry["clickFlag"] = cf
                    tac = e_row.get("textAfterClick")
                    if tac:
                        entry["textAfterClick"] = tac
                    entries.append(entry)

            if entries:
                desc["eventEntries"] = entries

            # 自动触发
            auto_key = desc_row.get("autoTriggerEventKey")
            if auto_key:
                desc["autoTriggerEventKey"] = auto_key

            # removeAfterInteraction
            if parse_bool(desc_row.get("removeAfterInteraction")):
                desc["removeAfterInteraction"] = True

            # seenFlag / viewLimit
            sf = desc_row.get("seenFlag")
            if sf:
                desc["seenFlag"] = sf
            vl = parse_int(desc_row.get("viewLimit"))
            if vl is not None:
                desc["viewLimit"] = vl

            # 时间/天气/季节限制
            tof = desc_row.get("timeOfDayRestriction")
            if tof:
                desc["timeOfDayRestriction"] = tof
            wr = desc_row.get("weatherRestriction")
            if wr:
                desc["weatherRestriction"] = wr
            sr = desc_row.get("seasonRestriction")
            if sr:
                desc["seasonRestriction"] = sr

            descriptions.append(desc)

        # 构建交互列表
        interaction_ids = parse_comma_list(s.get("interactions"))
        interactions = []
        for iid in interaction_ids:
            matched = interaction_map.get(scene_id, [])
            # 通过name匹配
            matched_interactions = [i for i in matched if i.get("name") == iid]
            if not matched_interactions:
                # 尝试通过costs等字段匹配
                matched_interactions = matched
            for i_row in matched_interactions:
                interaction = build_interaction(i_row, cost_map, bp_map, conditions)
                if interaction:
                    interactions.append(interaction)

        # 构建基础场景对象
        base = {
            "id": scene_id,
            "name": s.get("name") or "",
            "notes": s.get("notes") or None,
            "descriptions": descriptions,
            "temperatureModifier": parse_int(s.get("temperatureModifier"), 0),
            "interactions": interactions,
            "isDungeon": parse_bool(s.get("isDungeon")),
        }

        # 背景图
        bg = s.get("backgroundImage")
        if bg:
            base["backgroundImage"] = bg

        # BGM
        bgm = s.get("bgmId")
        if bgm:
            base["bgmId"] = bgm

        if is_sub:
            # 子场景
            sub = {
                **base,
                "parentSceneId": s.get("parentSceneId") or "",
            }
            # 相邻子场景（地牢用）
            adj = s.get("adjacentSubScenes")
            if adj:
                parts = parse_comma_list(adj)
                if len(parts) >= 6:
                    dirs = ["north", "south", "east", "west", "up", "down"]
                    adj_dict = {}
                    for i, d in enumerate(dirs):
                        if parts[i] and parts[i].lower() != "none":
                            adj_dict[d] = parts[i]
                    if adj_dict:
                        sub["adjacentSubScenes"] = adj_dict
            sub_scenes[scene_id] = sub
        else:
            # 主场景
            sub_ids = s.get("subSceneIds")
            if sub_ids:
                parts = parse_comma_list(sub_ids)
                if parts:
                    base["subSceneIds"] = parts
            scenes[scene_id] = base

    return {"scenes": scenes, "subScenes": sub_scenes}


def build_interaction(i_row: dict, cost_map: dict, bp_map: dict, conditions: dict) -> dict | None:
    """构建单个交互对象"""
    iname = i_row.get("name")
    if not iname:
        return None

    interaction = {
        "id": iname,  # 使用name作为id
        "name": iname,
        "interactionType": i_row.get("interactionType") or "EXPLORE",
        "displayPriority": parse_int(i_row.get("displayPriority"), 0),
        "isOneTime": parse_bool(i_row.get("isOneTime")),
        "cooldownMinutes": parse_int(i_row.get("cooldownMinutes"), 0),
    }

    # 描述
    desc = i_row.get("description")
    if desc:
        interaction["description"] = desc

    # 显示/可用条件
    dcond = i_row.get("displayCondition")
    if dcond:
        interaction["displayCondition"] = dcond
    acond = i_row.get("availableCondition")
    if acond:
        interaction["availableCondition"] = acond
    ut = i_row.get("unavailableTooltip")
    if ut:
        interaction["unavailableTooltip"] = ut

    # 消耗
    cost_ids = parse_comma_list(i_row.get("costs"))
    costs = []
    for cid in cost_ids:
        cost_row = cost_map.get(cid)
        if cost_row:
            cost = {
                "costType": cost_row.get("costType") or "STAMINA",
                "value": parse_int(cost_row.get("value"), 0),
                "affectedByCoefficient": parse_bool(cost_row.get("affectedByCoefficient")),
            }
            item_id = cost_row.get("itemId")
            if item_id:
                cost["itemId"] = item_id
            item_qty = parse_int(cost_row.get("itemQuantity"))
            if item_qty is not None:
                cost["itemQuantity"] = item_qty
            costs.append(cost)
    if costs:
        interaction["costs"] = costs

    # 确认弹窗
    if parse_bool(i_row.get("requiresConfirmation")):
        interaction["requiresConfirmation"] = True
        ct = i_row.get("confirmationText")
        if ct:
            interaction["confirmationText"] = ct

    # usedFlag
    uf = i_row.get("usedFlag")
    if uf:
        interaction["usedFlag"] = uf
    cfp = i_row.get("cooldownFlagPrefix")
    if cfp:
        interaction["cooldownFlagPrefix"] = cfp

    # 行为参数
    bp_ids = parse_comma_list(i_row.get("behaviorParams"))
    behavior_params = None
    for bid in bp_ids:
        bp_row = bp_map.get(bid)
        if bp_row:
            bp_type = bp_row.get("interactionType") or interaction["interactionType"]
            behavior_params = build_behavior_params(bp_type, bp_row)
            if behavior_params:
                break

    if behavior_params is None:
        behavior_params = {"interactionType": interaction["interactionType"]}
    interaction["behaviorParams"] = behavior_params

    return interaction


def build_behavior_params(bp_type: str, bp_row: dict) -> dict:
    """构建交互行为参数"""
    params = {"interactionType": bp_type}

    bp_type_upper = bp_type.upper()

    if bp_type_upper == "EXPLORE":
        pass
    elif bp_type_upper == "EVENT":
        ev = bp_row.get("eventId")
        if ev:
            params["eventId"] = ev
    elif bp_type_upper == "FUNCTION":
        ft = bp_row.get("functionType")
        if ft:
            params["functionType"] = ft
    elif bp_type_upper == "ENTER_SUB_SCENE":
        ss = bp_row.get("subSceneId")
        if ss:
            params["subSceneId"] = ss
    elif bp_type_upper == "EXIT_SUB_SCENE":
        pass
    elif bp_type_upper == "MOVE":
        d = bp_row.get("direction")
        if d:
            params["direction"] = d
    elif bp_type_upper == "REST":
        pass
    elif bp_type_upper == "TALK":
        ev = bp_row.get("eventId")
        if ev:
            params["eventId"] = ev
    elif bp_type_upper == "TRADE":
        tr = bp_row.get("traderId")
        if tr:
            params["traderId"] = tr
    elif bp_type_upper == "MOVE_TO_SCENE":
        ts = bp_row.get("targetSceneId")
        if ts:
            params["targetSceneId"] = ts
        tn = bp_row.get("targetNodeId")
        if tn:
            params["targetNodeId"] = tn
        tt = parse_int(bp_row.get("travelTimeMinutes"))
        if tt is not None:
            params["travelTimeMinutes"] = tt
        sc = parse_int(bp_row.get("staminaCost"))
        if sc is not None:
            params["staminaCost"] = sc
        pd = bp_row.get("pathDescription")
        if pd:
            params["pathDescription"] = pd
        ep = bp_row.get("encounterEventPool")
        if ep:
            # 解析 encounterEventPool 格式: "eventId:weight,eventId:weight"
            pool = []
            for item in parse_comma_list(ep):
                parts = item.split(":")
                if len(parts) >= 2:
                    pool.append({"eventId": parts[0].strip(), "weight": int(parts[1].strip())})
            if pool:
                params["encounterEventPool"] = pool
        tss = bp_row.get("targetSubSceneId")
        if tss:
            params["targetSubSceneId"] = tss
        tmi = bp_row.get("targetMapId")
        if tmi:
            params["targetMapId"] = tmi
        req = bp_row.get("requirements")
        if req:
            params["requirements"] = req

    return params


# ============================================================
# TypeScript 代码生成器
# ============================================================

def generate_ts(config: dict, conditions: dict) -> str:
    """生成 TypeScript 源码字符串"""
    lines = []
    indent = "  "

    # 头部导入
    lines.append("// ============================================================")
    lines.append("// 此文件由 XLSX 转换器自动生成，请勿手动修改")
    lines.append("// 生成时间: " + __import__("datetime").datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    lines.append("// ============================================================")
    lines.append("")
    lines.append("import {")
    lines.append("  AttributeType,")
    lines.append("  ComparisonOperator,")
    lines.append("  ConditionTargetType,")
    lines.append("  LogicOperator,")
    lines.append("} from '@/types/effect'")
    lines.append("import type { Scene, SubScene, SceneRegistry } from '../types/scene'")
    lines.append("import {")
    lines.append("  InteractionType, InteractionCostType, Direction,")
    lines.append("} from '../types/scene'")
    lines.append("")

    # 生成条件常量（按拓扑排序：简单条件先于复合条件）
    if conditions:
        lines.append("// ============================================================")
        lines.append("// 条件定义")
        lines.append("// ============================================================")
        lines.append("")

        # 拓扑排序：简单条件（无 subConditions）先输出
        simple_conds = {cid: c for cid, c in conditions.items() if "subConditions" not in c}
        compound_conds = {cid: c for cid, c in conditions.items() if "subConditions" in c}

        for cid, cond in simple_conds.items():
            lines.append(f"const {cid} = {condition_to_ts(cond, indent)}")
        for cid, cond in compound_conds.items():
            lines.append(f"const {cid} = {condition_to_ts(cond, indent)}")
        lines.append("")

    # 生成场景
    scenes = config.get("scenes", {})
    sub_scenes = config.get("subScenes", {})

    for sid, scene in scenes.items():
        lines.append("// ============================================================")
        lines.append(f"// {scene.get('name', sid)}")
        notes = scene.get("notes")
        if notes:
            lines.append(f"// {notes}")
        lines.append("// ============================================================")
        lines.append("")
        lines.append(f"const {sid}: Scene = {scene_to_ts(scene, conditions, indent)}")
        lines.append("")

    for sid, sub in sub_scenes.items():
        lines.append("// ============================================================")
        lines.append(f"// {sub.get('name', sid)} (子场景)")
        notes = sub.get("notes")
        if notes:
            lines.append(f"// {notes}")
        lines.append("// ============================================================")
        lines.append("")
        lines.append(f"const {sid}: SubScene = {scene_to_ts(sub, conditions, indent)}")
        lines.append("")

    # 生成注册表
    all_scene_ids = [f"  {sid}," for sid in scenes.keys()]
    all_sub_scene_ids = [f"  {sid}," for sid in sub_scenes.keys()]

    # 找到初始场景
    initial_scene = "beach" if "beach" in scenes else (list(scenes.keys())[0] if scenes else "")

    lines.append("// ============================================================")
    lines.append("// 场景注册表")
    lines.append("// ============================================================")
    lines.append("")
    lines.append("export const sceneRegistry: SceneRegistry = {")
    lines.append("  scenes: {")
    for sid in scenes.keys():
        lines.append(f"    {sid},")
    lines.append("  },")
    lines.append("  subScenes: {")
    for sid in sub_scenes.keys():
        lines.append(f"    {sid},")
    lines.append("  },")
    lines.append(f"  initialSceneId: '{initial_scene}',")
    lines.append("}")
    lines.append("")

    return "\n".join(lines)


def condition_to_ts(cond: dict, indent: str = "  ") -> str:
    """将条件字典转换为 TypeScript 对象字符串"""
    lines = ["{"]
    inner_indent = indent + "  "

    for key, val in cond.items():
        if key == "subConditions":
            sub_ids = val
            lines.append(f"{inner_indent}{key}: [")
            for sub_id in sub_ids:
                lines.append(f"{inner_indent}  {sub_id},")
            lines.append(f"{inner_indent}],")
        elif key == "target":
            target = val
            lines.append(f"{inner_indent}{key}: {{")
            for tk, tv in target.items():
                if tv is not None:
                    if tk == "type":
                        lines.append(f"{inner_indent}  type: {to_condition_target_type(tv)},")
                    elif tk == "attributeType":
                        lines.append(f"{inner_indent}  attributeType: {to_attribute_type(tv)},")
                    else:
                        lines.append(f"{inner_indent}  {tk}: {to_ts_value(tv)},")
            lines.append(f"{inner_indent}}},")
        elif key == "operator":
            op = str(val).upper() if val else None
            if op:
                if op == "EQUAL":
                    lines.append(f"{inner_indent}{key}: ComparisonOperator.EQUAL,")
                elif op == "NOT_EQUAL":
                    lines.append(f"{inner_indent}{key}: ComparisonOperator.NOT_EQUAL,")
                elif op == "GREATER":
                    lines.append(f"{inner_indent}{key}: ComparisonOperator.GREATER,")
                elif op == "GREATER_EQUAL":
                    lines.append(f"{inner_indent}{key}: ComparisonOperator.GREATER_EQUAL,")
                elif op == "LESS":
                    lines.append(f"{inner_indent}{key}: ComparisonOperator.LESS,")
                elif op == "LESS_EQUAL":
                    lines.append(f"{inner_indent}{key}: ComparisonOperator.LESS_EQUAL,")
                elif op == "BETWEEN":
                    lines.append(f"{inner_indent}{key}: ComparisonOperator.BETWEEN,")
                elif op == "IN":
                    lines.append(f"{inner_indent}{key}: ComparisonOperator.IN,")
                elif op == "NOT_IN":
                    lines.append(f"{inner_indent}{key}: ComparisonOperator.NOT_IN,")
                elif op == "EXISTS":
                    lines.append(f"{inner_indent}{key}: ComparisonOperator.EXISTS,")
                elif op == "NOT_EXISTS":
                    lines.append(f"{inner_indent}{key}: ComparisonOperator.NOT_EXISTS,")
                else:
                    lines.append(f"{inner_indent}{key}: '{val}',")
        elif key == "logic":
            logic = str(val).upper() if val else None
            if logic == "AND":
                lines.append(f"{inner_indent}{key}: LogicOperator.AND,")
            elif logic == "OR":
                lines.append(f"{inner_indent}{key}: LogicOperator.OR,")
            elif logic == "NOT":
                lines.append(f"{inner_indent}{key}: LogicOperator.NOT,")
            else:
                lines.append(f"{inner_indent}{key}: '{val}',")
        elif key == "value":
            if val is not None:
                lines.append(f"{inner_indent}{key}: {to_ts_value(val)},")
        elif key == "value2":
            if val is not None:
                lines.append(f"{inner_indent}{key}: {to_ts_value(val)},")
        else:
            if val is not None:
                lines.append(f"{inner_indent}{key}: {to_ts_value(val)},")

    lines.append(f"{indent}}}")
    return "\n".join(lines)


def to_condition_target_type(t: str) -> str:
    """将条件目标类型字符串转换为 ConditionTargetType 枚举"""
    mapping = {
        "ATTRIBUTE": "ConditionTargetType.ATTRIBUTE",
        "FLAG": "ConditionTargetType.FLAG",
        "ITEM": "ConditionTargetType.ITEM",
        "STATUS": "ConditionTargetType.STATUS",
        "SCENE": "ConditionTargetType.SCENE",
        "TIME": "ConditionTargetType.TIME",
        "WEATHER": "ConditionTargetType.WEATHER",
        "SEASON": "ConditionTargetType.SEASON",
        "SAN_LEVEL": "ConditionTargetType.SAN_LEVEL",
        "CORRUPTION": "ConditionTargetType.CORRUPTION",
        "SKILL": "ConditionTargetType.SKILL",
        "WEAPON_PROFICIENCY": "ConditionTargetType.WEAPON_PROFICIENCY",
        "RECIPE_UNLOCKED": "ConditionTargetType.RECIPE_UNLOCKED",
        "PLAYER_GOLD": "ConditionTargetType.PLAYER_GOLD",
        "CARRY_WEIGHT_RATE": "ConditionTargetType.CARRY_WEIGHT_RATE",
    }
    return mapping.get(t.upper(), f"'{t}'")


def to_attribute_type(t: str) -> str:
    """将属性类型字符串转换为 AttributeType 枚举"""
    mapping = {
        "HP": "AttributeType.HP",
        "SATIETY": "AttributeType.SATIETY",
        "STAMINA": "AttributeType.STAMINA",
        "SAN": "AttributeType.SAN",
        "WARMTH": "AttributeType.WARMTH",
        "CARRY_WEIGHT": "AttributeType.CARRY_WEIGHT",
        "STRENGTH": "AttributeType.STRENGTH",
        "AGILITY": "AttributeType.AGILITY",
        "INTELLIGENCE": "AttributeType.INTELLIGENCE",
        "CONSTITUTION": "AttributeType.CONSTITUTION",
        "STRENGTH_EXP": "AttributeType.STRENGTH_EXP",
        "AGILITY_EXP": "AttributeType.AGILITY_EXP",
        "INTELLIGENCE_EXP": "AttributeType.INTELLIGENCE_EXP",
        "CONSTITUTION_EXP": "AttributeType.CONSTITUTION_EXP",
        "WEAPON_PROFICIENCY": "AttributeType.WEAPON_PROFICIENCY",
        "WEAPON_PROFICIENCY_EXP": "AttributeType.WEAPON_PROFICIENCY_EXP",
        "SLASH_DEFENSE": "AttributeType.SLASH_DEFENSE",
        "BLUNT_DEFENSE": "AttributeType.BLUNT_DEFENSE",
        "RANGED_DEFENSE": "AttributeType.RANGED_DEFENSE",
        "POISON_DEFENSE": "AttributeType.POISON_DEFENSE",
        "FIRE_DEFENSE": "AttributeType.FIRE_DEFENSE",
        "SKILL_LEVEL": "AttributeType.SKILL_LEVEL",
        "SKILL_EXP": "AttributeType.SKILL_EXP",
        "RECOVERY_RATE_COEFFICIENT": "AttributeType.RECOVERY_RATE_COEFFICIENT",
        "SATIETY_UPPER_LIMIT_COEFFICIENT": "AttributeType.SATIETY_UPPER_LIMIT_COEFFICIENT",
        "SATIETY_LOSS_COEFFICIENT": "AttributeType.SATIETY_LOSS_COEFFICIENT",
        "STAMINA_CONSUMPTION_COEFFICIENT": "AttributeType.STAMINA_CONSUMPTION_COEFFICIENT",
        "STAMINA_RECOVERY_COEFFICIENT": "AttributeType.STAMINA_RECOVERY_COEFFICIENT",
        "STAMINA_RECOVERY_FIX": "AttributeType.STAMINA_RECOVERY_FIX",
        "SAN_MODIFIER": "AttributeType.SAN_MODIFIER",
        "TEMPERATURE_LOW": "AttributeType.TEMPERATURE_LOW",
        "TEMPERATURE_HIGH": "AttributeType.TEMPERATURE_HIGH",
        "CARRY_WEIGHT_MODIFIER": "AttributeType.CARRY_WEIGHT_MODIFIER",
    }
    return mapping.get(t.upper(), f"'{t}'")


def scene_to_ts(scene: dict, conditions: dict, indent: str = "  ") -> str:
    """将场景字典转换为 TypeScript 对象字符串"""
    lines = ["{"]
    inner = indent + "  "

    # id
    lines.append(f"{inner}id: '{scene.get('id', '')}',")
    # name
    lines.append(f"{inner}name: '{scene.get('name', '')}',")

    notes = scene.get("notes")
    if notes:
        lines.append(f"{inner}notes: '{notes}',")

    # descriptions
    descs = scene.get("descriptions", [])
    lines.append(f"{inner}descriptions: [")
    for d in descs:
        lines.append(f"{inner}  {description_to_ts(d, conditions, inner + '  ')},")
    lines.append(f"{inner}],")

    # backgroundImage
    bg = scene.get("backgroundImage")
    if bg:
        lines.append(f"{inner}backgroundImage: '{bg}',")

    # temperatureModifier
    lines.append(f"{inner}temperatureModifier: {scene.get('temperatureModifier', 0)},")

    # interactions
    interactions = scene.get("interactions", [])
    lines.append(f"{inner}interactions: [")
    for i in interactions:
        lines.append(f"{inner}  {interaction_to_ts(i, conditions, inner + '  ')},")
    lines.append(f"{inner}],")

    # isDungeon
    lines.append(f"{inner}isDungeon: {str(scene.get('isDungeon', False)).lower()},")

    # 子场景特有字段
    psid = scene.get("parentSceneId")
    if psid:
        lines.append(f"{inner}parentSceneId: '{psid}',")

    # subSceneIds
    sub_ids = scene.get("subSceneIds")
    if sub_ids:
        ids_str = ", ".join(f"'{sid}'" for sid in sub_ids)
        lines.append(f"{inner}subSceneIds: [{ids_str}],")

    # adjacentSubScenes
    adj = scene.get("adjacentSubScenes")
    if adj:
        lines.append(f"{inner}adjacentSubScenes: {{")
        for d, sid in adj.items():
            lines.append(f"{inner}  {d}: '{sid}',")
        lines.append(f"{inner}}},")

    # bgmId
    bgm = scene.get("bgmId")
    if bgm:
        lines.append(f"{inner}bgmId: '{bgm}',")

    lines.append(f"{indent}}}")
    return "\n".join(lines)


def description_to_ts(desc: dict, conditions: dict, indent: str) -> str:
    """将描述字典转换为 TypeScript 对象字符串"""
    lines = ["{"]
    inner = indent + "  "

    lines.append(f"{inner}id: '{desc.get('id', '')}',")
    lines.append(f"{inner}priority: {desc.get('priority', 0)},")

    # 权重（可选）
    w = desc.get("weight")
    if w is not None:
        lines.append(f"{inner}weight: {w},")

    # 显示条件
    dc = desc.get("displayCondition")
    if dc:
        if isinstance(dc, str) and dc in conditions:
            lines.append(f"{inner}displayCondition: {dc},")
        else:
            lines.append(f"{inner}displayCondition: {dc},")

    # 文本
    text = desc.get("text", "")
    if text:
        lines.append(f"{inner}text: {to_ts_value(text)},")

    # 事件入口
    entries = desc.get("eventEntries")
    if entries:
        lines.append(f"{inner}eventEntries: [")
        for e in entries:
            lines.append(f"{inner}  {entry_to_ts(e, conditions, inner + '  ')},")
        lines.append(f"{inner}],")

    # isAutoTrigger
    lines.append(f"{inner}isAutoTrigger: {str(desc.get('isAutoTrigger', False)).lower()},")

    # autoTriggerEventKey
    atk = desc.get("autoTriggerEventKey")
    if atk:
        lines.append(f"{inner}autoTriggerEventKey: '{atk}',")

    # removeAfterInteraction
    if desc.get("removeAfterInteraction"):
        lines.append(f"{inner}removeAfterInteraction: true,")

    # isOneTime
    lines.append(f"{inner}isOneTime: {str(desc.get('isOneTime', False)).lower()},")

    # seenFlag
    sf = desc.get("seenFlag")
    if sf:
        lines.append(f"{inner}seenFlag: '{sf}',")

    # viewLimit
    vl = desc.get("viewLimit")
    if vl is not None:
        lines.append(f"{inner}viewLimit: {vl},")

    lines.append(f"{indent}}}")
    return "\n".join(lines)


def entry_to_ts(entry: dict, conditions: dict, indent: str) -> str:
    """将事件入口字典转换为 TypeScript 对象字符串"""
    lines = ["{"]
    inner = indent + "  "

    lines.append(f"{inner}key: '{entry.get('key', '')}',")
    lines.append(f"{inner}displayText: '{entry.get('displayText', '')}',")
    lines.append(f"{inner}eventId: '{entry.get('eventId', '')}',")

    dc = entry.get("displayCondition")
    if dc:
        if isinstance(dc, str) and dc in conditions:
            lines.append(f"{inner}displayCondition: {dc},")
        else:
            lines.append(f"{inner}displayCondition: {dc},")

    ac = entry.get("availableCondition")
    if ac:
        if isinstance(ac, str) and ac in conditions:
            lines.append(f"{inner}availableCondition: {ac},")
        else:
            lines.append(f"{inner}availableCondition: {ac},")

    if entry.get("removeAfterClick"):
        lines.append(f"{inner}removeAfterClick: true,")

    cf = entry.get("clickFlag")
    if cf:
        lines.append(f"{inner}clickFlag: '{cf}',")

    tac = entry.get("textAfterClick")
    if tac:
        lines.append(f"{inner}textAfterClick: '{tac}',")

    lines.append(f"{indent}}}")
    return "\n".join(lines)


def interaction_to_ts(interaction: dict, conditions: dict, indent: str) -> str:
    """将交互按钮字典转换为 TypeScript 对象字符串"""
    lines = ["{"]
    inner = indent + "  "

    lines.append(f"{inner}id: '{interaction.get('id', '')}',")
    lines.append(f"{inner}name: '{interaction.get('name', '')}',")

    desc = interaction.get("description")
    if desc:
        lines.append(f"{inner}description: '{desc}',")

    # displayCondition
    dc = interaction.get("displayCondition")
    if dc:
        if isinstance(dc, str) and dc in conditions:
            lines.append(f"{inner}displayCondition: {dc},")
        else:
            lines.append(f"{inner}displayCondition: {dc},")

    # availableCondition
    ac = interaction.get("availableCondition")
    if ac:
        if isinstance(ac, str) and ac in conditions:
            lines.append(f"{inner}availableCondition: {ac},")
        else:
            lines.append(f"{inner}availableCondition: {ac},")

    ut = interaction.get("unavailableTooltip")
    if ut:
        lines.append(f"{inner}unavailableTooltip: '{ut}',")

    # interactionType
    itype = interaction.get("interactionType", "EXPLORE")
    if itype.startswith("InteractionType."):
        lines.append(f"{inner}interactionType: {itype},")
    else:
        lines.append(f"{inner}interactionType: InteractionType.{itype.upper()},")

    # costs
    costs = interaction.get("costs")
    if costs:
        lines.append(f"{inner}costs: [")
        for c in costs:
            lines.append(f"{inner}  {cost_to_ts(c, inner + '  ')},")
        lines.append(f"{inner}],")

    # behaviorParams
    bp = interaction.get("behaviorParams", {})
    lines.append(f"{inner}behaviorParams: {behavior_params_to_ts(bp, conditions, inner)},")

    # requiresConfirmation
    if interaction.get("requiresConfirmation"):
        lines.append(f"{inner}requiresConfirmation: true,")
        ct = interaction.get("confirmationText")
        if ct:
            lines.append(f"{inner}confirmationText: '{ct}',")

    lines.append(f"{inner}displayPriority: {interaction.get('displayPriority', 0)},")
    lines.append(f"{inner}isOneTime: {str(interaction.get('isOneTime', False)).lower()},")

    uf = interaction.get("usedFlag")
    if uf:
        lines.append(f"{inner}usedFlag: '{uf}',")

    lines.append(f"{inner}cooldownMinutes: {interaction.get('cooldownMinutes', 0)},")

    cfp = interaction.get("cooldownFlagPrefix")
    if cfp:
        lines.append(f"{inner}cooldownFlagPrefix: '{cfp}',")

    lines.append(f"{indent}}}")
    return "\n".join(lines)


def cost_to_ts(cost: dict, indent: str) -> str:
    """将花费字典转换为 TypeScript 对象字符串"""
    lines = ["{"]
    inner = indent + "  "

    ct = cost.get("costType", "STAMINA")
    if ct.startswith("InteractionCostType."):
        lines.append(f"{inner}costType: {ct},")
    else:
        lines.append(f"{inner}costType: InteractionCostType.{ct.upper()},")

    lines.append(f"{inner}value: {cost.get('value', 0)},")
    lines.append(f"{inner}affectedByCoefficient: {str(cost.get('affectedByCoefficient', False)).lower()},")

    item_id = cost.get("itemId")
    if item_id:
        lines.append(f"{inner}itemId: '{item_id}',")
    item_qty = cost.get("itemQuantity")
    if item_qty is not None:
        lines.append(f"{inner}itemQuantity: {item_qty},")

    lines.append(f"{indent}}}")
    return "\n".join(lines)


def behavior_params_to_ts(bp: dict, conditions: dict, indent: str) -> str:
    """将行为参数字典转换为 TypeScript 对象字符串"""
    lines = ["{"]
    inner = indent + "  "

    bp_type = bp.get("interactionType", "EXPLORE")
    if bp_type.startswith("InteractionType."):
        lines.append(f"{inner}interactionType: {bp_type},")
    else:
        lines.append(f"{inner}interactionType: InteractionType.{bp_type.upper()},")

    # 根据类型添加特定字段
    bp_type_upper = bp_type.upper() if not bp_type.startswith("InteractionType.") else bp_type.split(".")[-1]

    ev = bp.get("eventId")
    if ev:
        ev_str = ", ".join(f"'{e}'" for e in (ev if isinstance(ev, list) else [ev]))
        if isinstance(ev, list) and len(ev) > 1:
            lines.append(f"{inner}eventId: [{ev_str}],")
        else:
            lines.append(f"{inner}eventId: '{ev}',")

    ft = bp.get("functionType")
    if ft:
        lines.append(f"{inner}functionType: FunctionType.{ft.upper()},")

    ss = bp.get("subSceneId")
    if ss:
        lines.append(f"{inner}subSceneId: '{ss}',")

    d = bp.get("direction")
    if d:
        dir_str = d.upper() if not d.startswith("Direction.") else d
        if dir_str.startswith("Direction."):
            lines.append(f"{inner}direction: {dir_str},")
        else:
            lines.append(f"{inner}direction: Direction.{dir_str},")

    ts = bp.get("targetSceneId")
    if ts:
        lines.append(f"{inner}targetSceneId: '{ts}',")

    tss = bp.get("targetSubSceneId")
    if tss:
        lines.append(f"{inner}targetSubSceneId: '{tss}',")

    tmi = bp.get("targetMapId")
    if tmi:
        lines.append(f"{inner}targetMapId: '{tmi}',")

    tn = bp.get("targetNodeId")
    if tn:
        lines.append(f"{inner}targetNodeId: '{tn}',")

    tt = bp.get("travelTimeMinutes")
    if tt is not None:
        lines.append(f"{inner}travelTimeMinutes: {tt},")

    sc = bp.get("staminaCost")
    if sc is not None:
        lines.append(f"{inner}staminaCost: {sc},")

    pd = bp.get("pathDescription")
    if pd:
        lines.append(f"{inner}pathDescription: '{pd}',")

    ep = bp.get("encounterEventPool")
    if ep:
        lines.append(f"{inner}encounterEventPool: [")
        for p_entry in ep:
            eid = p_entry.get("eventId", "")
            w = p_entry.get("weight", 1)
            cond = p_entry.get("condition")
            if cond and isinstance(cond, str) and cond in conditions:
                lines.append(f"{inner}  {{ eventId: '{eid}', weight: {w}, condition: {cond} }},")
            else:
                lines.append(f"{inner}  {{ eventId: '{eid}', weight: {w} }},")
        lines.append(f"{inner}],")

    req = bp.get("requirements")
    if req:
        lines.append(f"{inner}requirements: {req},")

    tr = bp.get("traderId")
    if tr:
        lines.append(f"{inner}traderId: '{tr}',")

    # weatherImpactCoefficient
    wic = bp.get("weatherImpactCoefficient")
    if wic is not None:
        lines.append(f"{inner}weatherImpactCoefficient: {wic},")

    lines.append(f"{indent}}}")
    return "\n".join(lines)


# ============================================================
# 主程序
# ============================================================

def main():
    print("=" * 60)
    print("场景 XLSX → TS 配置转换器")
    print("=" * 60)

    # 验证文件存在
    for fpath, fname in [(SCENE_XLSX, "场景.xlsx"), (COND_XLSX, "条件.xlsx")]:
        if not os.path.exists(fpath):
            print(f"错误: 找不到 {fname} 文件: {fpath}")
            sys.exit(1)
        print(f"✓ 已找到 {fname}")

    # 1. 解析条件
    print("\n▶ 正在解析条件...")
    conditions = parse_conditions(COND_XLSX)
    print(f"  ✓ 解析了 {len(conditions)} 个条件")
    for cid, cond in conditions.items():
        if "subConditions" in cond:
            print(f"    - {cid}: 复合条件 (logic={cond.get('logic')}, {len(cond['subConditions'])} 个子条件)")
        else:
            print(f"    - {cid}: 简单条件 (target={cond.get('target', {}).get('type')}, operator={cond.get('operator')})")

    # 2. 构建场景配置
    print("\n▶ 正在解析场景...")
    config = build_scene_config(SCENE_XLSX, conditions)
    scenes = config.get("scenes", {})
    sub_scenes = config.get("subScenes", {})
    print(f"  ✓ 解析了 {len(scenes)} 个主场景")
    for sid, scene in scenes.items():
        desc_count = len(scene.get("descriptions", []))
        inter_count = len(scene.get("interactions", []))
        print(f"    - {sid}: {desc_count} 个描述, {inter_count} 个交互")
    print(f"  ✓ 解析了 {len(sub_scenes)} 个子场景")
    for sid, sub in sub_scenes.items():
        desc_count = len(sub.get("descriptions", []))
        inter_count = len(sub.get("interactions", []))
        print(f"    - {sid}: {desc_count} 个描述, {inter_count} 个交互")

    # 3. 生成 TypeScript 代码
    print("\n▶ 正在生成 TypeScript 代码...")
    ts_code = generate_ts(config, conditions)

    # 4. 写入文件
    os.makedirs(os.path.dirname(OUTPUT_TS), exist_ok=True)
    with open(OUTPUT_TS, "w", encoding="utf-8") as f:
        f.write(ts_code)
    print(f"  ✓ 文件已写入: {OUTPUT_TS}")
    print(f"  ✓ 共 {len(ts_code.splitlines())} 行代码")

    # 5. 验证
    print("\n▶ 正在验证生成的数据...")
    errors = []
    for sid, scene in scenes.items():
        for desc in scene.get("descriptions", []):
            dc = desc.get("displayCondition")
            if dc and isinstance(dc, str) and dc not in conditions:
                errors.append(f"场景 '{sid}' 描述 '{desc.get('id')}' 引用了不存在的条件 '{dc}'")
            for entry in desc.get("eventEntries", []):
                edc = entry.get("displayCondition")
                if edc and isinstance(edc, str) and edc not in conditions:
                    errors.append(f"场景 '{sid}' 入口 '{entry.get('key')}' 引用了不存在的条件 '{edc}'")
                eac = entry.get("availableCondition")
                if eac and isinstance(eac, str) and eac not in conditions:
                    errors.append(f"场景 '{sid}' 入口 '{entry.get('key')}' 引用了不存在的条件 '{eac}'")

    for sid, sub in sub_scenes.items():
        for desc in sub.get("descriptions", []):
            dc = desc.get("displayCondition")
            if dc and isinstance(dc, str) and dc not in conditions:
                errors.append(f"子场景 '{sid}' 描述 '{desc.get('id')}' 引用了不存在的条件 '{dc}'")

    if errors:
        print(f"  ⚠ 发现 {len(errors)} 个问题:")
        for err in errors:
            print(f"    - {err}")
    else:
        print("  ✓ 所有条件引用均正确")

    print("\n" + "=" * 60)
    print("转换完成！")
    print("=" * 60)


if __name__ == "__main__":
    main()