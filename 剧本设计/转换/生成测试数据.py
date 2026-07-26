"""
生成场景.xlsx 和 条件.xlsx 的测试数据
严格遵循 types/scene.ts 和 types/effect.ts 中定义的数据结构
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from copy import copy


# ============================================================
# 通用工具函数
# ============================================================

def write_header(ws, headers: list[str]):
    """写入表头行并设置样式"""
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def write_row(ws, row_idx: int, values: list):
    """写入一行数据"""
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.alignment = Alignment(wrap_text=True, vertical="top")


# ============================================================
# 创建场景.xlsx
# ============================================================

scene_wb = openpyxl.Workbook()

# ---- 场景 Sheet ----
ws_scene = scene_wb.active
ws_scene.title = "场景"
write_header(ws_scene, [
    "id", "name\n名称", "descriptions\n描述列表(逗号分隔)",
    "temperatureModifier\n温度影响", "interactions\n固定交互按钮列表(逗号分隔)",
    "isDungeon\n是否地牢场景", "是否子场景", "parentSceneId\n母场景ID",
    "subSceneIds\n子场景列表(逗号分隔)",
    "adjacentSubScenes\n[north,south,east,west,up,down]"
])

scene_data = [
    # 海滩 - 主场景
    {
        "id": "beach", "name": "海滩", "descriptions": "beach_first_arrival,beach_second_arrival,beach_crabs",
        "temperatureModifier": 5, "interactions": "探索海滩,进入洞穴,前往森林,休息",
        "isDungeon": "False", "is_sub_scene": "False", "parentSceneId": None,
        "subSceneIds": "beach_cave", "adjacentSubScenes": None
    },
    # 森林 - 主场景
    {
        "id": "forest", "name": "森林", "descriptions": "forest_first,forest_normal",
        "temperatureModifier": -3, "interactions": "探索森林,返回海滩,休息",
        "isDungeon": "False", "is_sub_scene": "False", "parentSceneId": None,
        "subSceneIds": None, "adjacentSubScenes": None
    },
    # 海滩洞穴 - 子场景
    {
        "id": "beach_cave", "name": "海滩洞穴", "descriptions": "cave_first_enter,cave_normal",
        "temperatureModifier": -10, "interactions": "探索洞穴,离开洞穴",
        "isDungeon": "False", "is_sub_scene": "True", "parentSceneId": "beach",
        "subSceneIds": None, "adjacentSubScenes": None
    },
]

for row_idx, scene in enumerate(scene_data, 2):
    write_row(ws_scene, row_idx, [
        scene["id"], scene["name"], scene["descriptions"],
        scene["temperatureModifier"], scene["interactions"],
        scene["isDungeon"], scene["is_sub_scene"], scene["parentSceneId"],
        scene["subSceneIds"], scene["adjacentSubScenes"]
    ])

# 设置列宽
ws_scene.column_dimensions['A'].width = 16
ws_scene.column_dimensions['B'].width = 12
ws_scene.column_dimensions['C'].width = 30
ws_scene.column_dimensions['D'].width = 16
ws_scene.column_dimensions['E'].width = 30
ws_scene.column_dimensions['F'].width = 12
ws_scene.column_dimensions['G'].width = 12
ws_scene.column_dimensions['H'].width = 16
ws_scene.column_dimensions['I'].width = 20
ws_scene.column_dimensions['J'].width = 30

# ---- 场景描述或事件入口 Sheet ----
ws_desc = scene_wb.create_sheet("场景描述或事件入口")
write_header(ws_desc, [
    "ID(所属场景或子场景id)", "id(描述id)", "priority\n优先级",
    "text\n文本内容", "displayCondition\n显示条件(条件id)",
    "eventEntries\n事件入口配置(逗号分隔的key列表)",
    "isAutoTrigger\n是否自动触发", "autoTriggerEventKey\n自动触发的事件入口key",
    "removeAfterInteraction\n事件入口触发后清除此描述",
    "isOneTime\n只能被看到一次", "seenFlag\n标志位", "viewLimit\n上限次数"
])

# 注意：文本中的 \n 在 Python 字符串中就是 \n，会被 XLSX 正确存储
descriptions = [
    # 海滩描述
    {
        "parent_id": "beach", "id": "beach_first_arrival", "priority": 10,
        "text": "你从破碎的座椅残骸中睁开眼睛。\n\n海浪的声音先于疼痛到达——规律的白噪音，一遍遍冲刷着沙滩。然后是气味：航空燃油的刺鼻味道混合着海水的咸腥。你嘴里有血的味道。\n\n右肋在痛。你低头检查，一片瘀伤从肋骨蔓延到腰侧。钝痛，不是骨折。至少不是让你立刻死去的那种。\n\n飞机残骸散落在沙滩上，像被巨人随手折断的玩具。机翼插在浅滩里，尾翼挂在远处的棕榈树上，机身断成几截，其中一截还在冒烟。\n\n你试着站起来。腿还听使唤。\n\n你还活着。",
        "displayCondition": None, "eventEntries": None,
        "isAutoTrigger": "False", "autoTriggerEventKey": None,
        "removeAfterInteraction": None, "isOneTime": "True",
        "seenFlag": "seen_beach_first_arrival", "viewLimit": None
    },
    {
        "parent_id": "beach", "id": "beach_second_arrival", "priority": 9,
        "text": "你努力站起身来。\n\n{plane_wreckage}依然散落在不远处，海浪不断拍打着沙滩。\n\n几只海鸥从天空飞过。\n\n或许你应该过去看看还有没有其他的幸存者。\n\n或者，看看有没有能用得到的东西。",
        "displayCondition": None, "eventEntries": "plane_wreckage",
        "isAutoTrigger": "False", "autoTriggerEventKey": None,
        "removeAfterInteraction": "True", "isOneTime": "False",
        "seenFlag": "seen_beach_plane_wreckage", "viewLimit": None
    },
    {
        "parent_id": "beach", "id": "beach_crabs", "priority": 3,
        "text": "你在沙滩上发现了一些{mutated_crab}在徘徊，它们的甲壳在阳光下泛着诡异的光泽。",
        "displayCondition": "cond_beach_crabs", "eventEntries": "mutated_crab",
        "isAutoTrigger": "True", "autoTriggerEventKey": "mutated_crab",
        "removeAfterInteraction": None, "isOneTime": "False",
        "seenFlag": None, "viewLimit": -1
    },
    # 森林描述
    {
        "parent_id": "forest", "id": "forest_first", "priority": 10,
        "text": "你踏入了茂密的森林。高大的树木遮蔽了大部分阳光，地面上铺满了落叶。空气中弥漫着泥土和植物的气息。{strange_trees}的树干上似乎有异常的突起。",
        "displayCondition": None, "eventEntries": "strange_trees",
        "isAutoTrigger": "False", "autoTriggerEventKey": None,
        "removeAfterInteraction": None, "isOneTime": "True",
        "seenFlag": "seen_forest_first", "viewLimit": 1
    },
    {
        "parent_id": "forest", "id": "forest_normal", "priority": 5,
        "text": "森林中一片静谧，偶尔能听到不知名生物的叫声。你可以看到一些{berry_bushes}在灌木丛中，以及一些{fallen_branches}。",
        "displayCondition": None, "eventEntries": "berry_bushes,fallen_branches",
        "isAutoTrigger": "False", "autoTriggerEventKey": None,
        "removeAfterInteraction": None, "isOneTime": "False",
        "seenFlag": None, "viewLimit": -1
    },
    # 洞穴描述
    {
        "parent_id": "beach_cave", "id": "cave_first_enter", "priority": 10,
        "text": "洞穴内部昏暗潮湿，空气中弥漫着霉味。你的眼睛逐渐适应了黑暗，看到洞穴深处似乎有{glowing_moss}在发光。地上散落着一些{journal_fragment_cave}。",
        "displayCondition": "cond_cave_first_enter", "eventEntries": "glowing_moss,journal_fragment_cave",
        "isAutoTrigger": "False", "autoTriggerEventKey": None,
        "removeAfterInteraction": None, "isOneTime": "True",
        "seenFlag": "seen_cave_first_enter", "viewLimit": 1
    },
    {
        "parent_id": "beach_cave", "id": "cave_normal", "priority": 5,
        "text": "洞穴内依然昏暗，{glowing_moss2}提供着微弱的照明。墙壁上有些{strange_markings}。",
        "displayCondition": None, "eventEntries": "glowing_moss2,strange_markings",
        "isAutoTrigger": "False", "autoTriggerEventKey": None,
        "removeAfterInteraction": None, "isOneTime": "False",
        "seenFlag": None, "viewLimit": -1
    },
]

for row_idx, desc in enumerate(descriptions, 2):
    write_row(ws_desc, row_idx, [
        desc["parent_id"], desc["id"], desc["priority"],
        desc["text"], desc["displayCondition"], desc["eventEntries"],
        desc["isAutoTrigger"], desc["autoTriggerEventKey"],
        desc["removeAfterInteraction"], desc["isOneTime"],
        desc["seenFlag"], desc["viewLimit"]
    ])

ws_desc.column_dimensions['A'].width = 16
ws_desc.column_dimensions['B'].width = 24
ws_desc.column_dimensions['C'].width = 10
ws_desc.column_dimensions['D'].width = 60
ws_desc.column_dimensions['E'].width = 20
ws_desc.column_dimensions['F'].width = 30
ws_desc.column_dimensions['G'].width = 14
ws_desc.column_dimensions['H'].width = 20
ws_desc.column_dimensions['I'].width = 20
ws_desc.column_dimensions['J'].width = 12
ws_desc.column_dimensions['K'].width = 20
ws_desc.column_dimensions['L'].width = 12

# ---- 事件入口配置 Sheet ----
ws_entry = scene_wb.create_sheet("事件入口配置")
write_header(ws_entry, [
    "ID(所属场景描述id)", "key", "displayText", "eventId",
    "displayCondition\n显示条件(条件id)", "availableCondition\n可用条件(条件id)",
    "removeAfterClick\n点击后是否移除", "clickFlag\n标志位",
    "textAfterClick\n点击后的文本变体"
])

entries = [
    # 海滩2 的事件入口
    {"parent_id": "beach_second_arrival", "key": "plane_wreckage", "displayText": "飞机残骸",
     "eventId": "event_plane_wreckage", "displayCondition": None, "availableCondition": None,
     "removeAfterClick": "False", "clickFlag": None, "textAfterClick": None},
    # 海滩3 的事件入口
    {"parent_id": "beach_crabs", "key": "mutated_crab", "displayText": "变异蟹",
     "eventId": "event_beach_crab_encounter", "displayCondition": None, "availableCondition": None,
     "removeAfterClick": "False", "clickFlag": None, "textAfterClick": None},
    # 森林1 的事件入口
    {"parent_id": "forest_first", "key": "strange_trees", "displayText": "变异的树木",
     "eventId": "event_strange_trees", "displayCondition": None, "availableCondition": None,
     "removeAfterClick": "False", "clickFlag": None, "textAfterClick": None},
    # 森林2 的事件入口
    {"parent_id": "forest_normal", "key": "berry_bushes", "displayText": "浆果灌木",
     "eventId": "event_gather_berries", "displayCondition": None, "availableCondition": None,
     "removeAfterClick": "False", "clickFlag": None, "textAfterClick": None},
    {"parent_id": "forest_normal", "key": "fallen_branches", "displayText": "落枝",
     "eventId": "event_gather_wood", "displayCondition": None, "availableCondition": None,
     "removeAfterClick": "False", "clickFlag": None, "textAfterClick": None},
    # 洞穴1 的事件入口
    {"parent_id": "cave_first_enter", "key": "glowing_moss", "displayText": "发光的苔藓",
     "eventId": "event_glowing_moss", "displayCondition": None, "availableCondition": None,
     "removeAfterClick": "False", "clickFlag": None, "textAfterClick": None},
    {"parent_id": "cave_first_enter", "key": "journal_fragment_cave", "displayText": "发黄的纸页",
     "eventId": "event_journal_fragment", "displayCondition": None, "availableCondition": None,
     "removeAfterClick": "False", "clickFlag": None, "textAfterClick": None},
    # 洞穴2 的事件入口
    {"parent_id": "cave_normal", "key": "glowing_moss2", "displayText": "发光的苔藓",
     "eventId": "event_glowing_moss", "displayCondition": None, "availableCondition": None,
     "removeAfterClick": "True", "clickFlag": "clicked_glowing_moss", "textAfterClick": "被采过的苔藓"},
    {"parent_id": "cave_normal", "key": "strange_markings", "displayText": "奇怪的刻痕",
     "eventId": "event_cave_markings", "displayCondition": "cond_strange_markings", "availableCondition": None,
     "removeAfterClick": "False", "clickFlag": None, "textAfterClick": None},
]

for row_idx, entry in enumerate(entries, 2):
    write_row(ws_entry, row_idx, [
        entry["parent_id"], entry["key"], entry["displayText"], entry["eventId"],
        entry["displayCondition"], entry["availableCondition"],
        entry["removeAfterClick"], entry["clickFlag"], entry["textAfterClick"]
    ])

ws_entry.column_dimensions['A'].width = 24
ws_entry.column_dimensions['B'].width = 20
ws_entry.column_dimensions['C'].width = 16
ws_entry.column_dimensions['D'].width = 24
ws_entry.column_dimensions['E'].width = 20
ws_entry.column_dimensions['F'].width = 20
ws_entry.column_dimensions['G'].width = 14
ws_entry.column_dimensions['H'].width = 20
ws_entry.column_dimensions['I'].width = 20

# ---- 场景交互按钮 Sheet ----
ws_interaction = scene_wb.create_sheet("场景交互按钮")
write_header(ws_interaction, [
    "ID(所属场景或子场景id)", "name", "description", "interactionType\n交互类型",
    "displayCondition\n显示条件(条件id)", "availableCondition\n可用条件(条件id)",
    "unavailableTooltip\n不可用提示文本",
    "costs\n执行花费(逗号分隔的ID)",
    "behaviorParams\n交互行为参数(逗号分隔的ID)",
    "requiresConfirmation\n交互弹窗", "confirmationText\n弹窗文本",
    "displayPriority\n优先级", "isOneTime\n只一次", "usedFlag\n标志位",
    "cooldownMinutes\n冷却时间", "cooldownFlagPrefix\n冷却标志位前缀",
])

interactions = [
    # 海滩交互
    {"parent_id": "beach", "name": "探索海滩", "description": None, "interactionType": "EXPLORE",
     "displayCondition": None, "availableCondition": None, "unavailableTooltip": None,
     "costs": "cost_explore_beach", "behaviorParams": "bp_explore_beach",
     "requiresConfirmation": "False", "confirmationText": None, "displayPriority": 10,
     "isOneTime": "False", "usedFlag": None, "cooldownMinutes": 0, "cooldownFlagPrefix": None},
    {"parent_id": "beach", "name": "进入洞穴", "description": None, "interactionType": "ENTER_SUB_SCENE",
     "displayCondition": "cond_enter_cave", "availableCondition": None, "unavailableTooltip": "尚未发现洞穴入口",
     "costs": "cost_enter_cave", "behaviorParams": "bp_enter_cave",
     "requiresConfirmation": "False", "confirmationText": None, "displayPriority": 5,
     "isOneTime": "False", "usedFlag": None, "cooldownMinutes": 0, "cooldownFlagPrefix": None},
    {"parent_id": "beach", "name": "前往森林", "description": None, "interactionType": "MOVE_TO_SCENE",
     "displayCondition": None, "availableCondition": None, "unavailableTooltip": None,
     "costs": "cost_move_to_forest", "behaviorParams": "bp_move_to_forest",
     "requiresConfirmation": "False", "confirmationText": None, "displayPriority": 3,
     "isOneTime": "False", "usedFlag": None, "cooldownMinutes": 0, "cooldownFlagPrefix": None},
    {"parent_id": "beach", "name": "休息", "description": "在海滩上稍作休息", "interactionType": "REST",
     "displayCondition": None, "availableCondition": None, "unavailableTooltip": None,
     "costs": "cost_rest_beach", "behaviorParams": "bp_rest_beach",
     "requiresConfirmation": "True", "confirmationText": "确定要在海滩上休息1小时吗？", "displayPriority": 1,
     "isOneTime": "False", "usedFlag": None, "cooldownMinutes": 0, "cooldownFlagPrefix": None},
    # 森林交互
    {"parent_id": "forest", "name": "探索森林", "description": None, "interactionType": "EXPLORE",
     "displayCondition": None, "availableCondition": None, "unavailableTooltip": None,
     "costs": "cost_explore_forest", "behaviorParams": "bp_explore_forest",
     "requiresConfirmation": "False", "confirmationText": None, "displayPriority": 10,
     "isOneTime": "False", "usedFlag": None, "cooldownMinutes": 0, "cooldownFlagPrefix": None},
    {"parent_id": "forest", "name": "返回海滩", "description": None, "interactionType": "MOVE_TO_SCENE",
     "displayCondition": None, "availableCondition": None, "unavailableTooltip": None,
     "costs": "cost_return_to_beach", "behaviorParams": "bp_return_to_beach",
     "requiresConfirmation": "False", "confirmationText": None, "displayPriority": 3,
     "isOneTime": "False", "usedFlag": None, "cooldownMinutes": 0, "cooldownFlagPrefix": None},
    {"parent_id": "forest", "name": "休息", "description": "在森林中稍作休息", "interactionType": "REST",
     "displayCondition": None, "availableCondition": None, "unavailableTooltip": None,
     "costs": "cost_rest_forest", "behaviorParams": "bp_rest_forest",
     "requiresConfirmation": "True", "confirmationText": "确定要在森林中休息1小时吗？", "displayPriority": 1,
     "isOneTime": "False", "usedFlag": None, "cooldownMinutes": 0, "cooldownFlagPrefix": None},
    # 洞穴交互
    {"parent_id": "beach_cave", "name": "探索洞穴", "description": None, "interactionType": "EXPLORE",
     "displayCondition": None, "availableCondition": None, "unavailableTooltip": None,
     "costs": "cost_explore_cave", "behaviorParams": "bp_explore_cave",
     "requiresConfirmation": "False", "confirmationText": None, "displayPriority": 10,
     "isOneTime": "False", "usedFlag": None, "cooldownMinutes": 0, "cooldownFlagPrefix": None},
    {"parent_id": "beach_cave", "name": "离开洞穴", "description": None, "interactionType": "EXIT_SUB_SCENE",
     "displayCondition": None, "availableCondition": None, "unavailableTooltip": None,
     "costs": "cost_exit_cave", "behaviorParams": "bp_exit_cave",
     "requiresConfirmation": "False", "confirmationText": None, "displayPriority": 1,
     "isOneTime": "False", "usedFlag": None, "cooldownMinutes": 0, "cooldownFlagPrefix": None},
]

for row_idx, interaction in enumerate(interactions, 2):
    write_row(ws_interaction, row_idx, [
        interaction["parent_id"], interaction["name"], interaction["description"],
        interaction["interactionType"], interaction["displayCondition"],
        interaction["availableCondition"], interaction["unavailableTooltip"],
        interaction["costs"], interaction["behaviorParams"],
        interaction["requiresConfirmation"], interaction["confirmationText"],
        interaction["displayPriority"], interaction["isOneTime"],
        interaction["usedFlag"], interaction["cooldownMinutes"],
        interaction["cooldownFlagPrefix"]
    ])

ws_interaction.column_dimensions['A'].width = 16
ws_interaction.column_dimensions['B'].width = 14
ws_interaction.column_dimensions['C'].width = 20
ws_interaction.column_dimensions['D'].width = 18
ws_interaction.column_dimensions['E'].width = 20
ws_interaction.column_dimensions['F'].width = 20
ws_interaction.column_dimensions['G'].width = 18
ws_interaction.column_dimensions['H'].width = 24
ws_interaction.column_dimensions['I'].width = 24
ws_interaction.column_dimensions['J'].width = 14
ws_interaction.column_dimensions['K'].width = 30
ws_interaction.column_dimensions['L'].width = 12
ws_interaction.column_dimensions['M'].width = 10
ws_interaction.column_dimensions['N'].width = 14
ws_interaction.column_dimensions['O'].width = 14
ws_interaction.column_dimensions['P'].width = 18

# ---- 交互花费 Sheet ----
ws_cost = scene_wb.create_sheet("交互花费")
write_header(ws_cost, [
    "ID(所属交互按钮id)", "costType", "value", "affectedByCoefficient",
    "itemId", "itemQuantity"
])

costs = [
    {"parent_id": "cost_explore_beach", "costType": "STAMINA", "value": 10, "affectedByCoefficient": "True", "itemId": None, "itemQuantity": None},
    {"parent_id": "cost_enter_cave", "costType": "STAMINA", "value": 15, "affectedByCoefficient": "True", "itemId": None, "itemQuantity": None},
    {"parent_id": "cost_move_to_forest", "costType": "STAMINA", "value": 25, "affectedByCoefficient": "True", "itemId": None, "itemQuantity": None},
    {"parent_id": "cost_rest_beach", "costType": "STAMINA", "value": 0, "affectedByCoefficient": "False", "itemId": None, "itemQuantity": None},
    {"parent_id": "cost_explore_forest", "costType": "STAMINA", "value": 12, "affectedByCoefficient": "True", "itemId": None, "itemQuantity": None},
    {"parent_id": "cost_return_to_beach", "costType": "STAMINA", "value": 25, "affectedByCoefficient": "True", "itemId": None, "itemQuantity": None},
    {"parent_id": "cost_rest_forest", "costType": "STAMINA", "value": 0, "affectedByCoefficient": "False", "itemId": None, "itemQuantity": None},
    {"parent_id": "cost_explore_cave", "costType": "STAMINA", "value": 15, "affectedByCoefficient": "True", "itemId": None, "itemQuantity": None},
    {"parent_id": "cost_exit_cave", "costType": "STAMINA", "value": 5, "affectedByCoefficient": "False", "itemId": None, "itemQuantity": None},
]

for row_idx, cost in enumerate(costs, 2):
    write_row(ws_cost, row_idx, [
        cost["parent_id"], cost["costType"], cost["value"],
        cost["affectedByCoefficient"], cost["itemId"], cost["itemQuantity"]
    ])

ws_cost.column_dimensions['A'].width = 24
ws_cost.column_dimensions['B'].width = 14
ws_cost.column_dimensions['C'].width = 10
ws_cost.column_dimensions['D'].width = 18
ws_cost.column_dimensions['E'].width = 14
ws_cost.column_dimensions['F'].width = 14

# ---- 交互行为参数 Sheet ----
ws_bp = scene_wb.create_sheet("交互行为参数")
write_header(ws_bp, [
    "ID(所属交互按钮id)", "interactionType", "eventId", "functionType",
    "subSceneId", "direction", "targetSceneId", "targetSubSceneId",
    "targetMapId", "targetNodeId", "travelTimeMinutes", "staminaCost",
    "pathDescription", "encounterEventPool", "requirements", "traderId"
])

behavior_params = [
    # 海滩行为参数
    {"parent_id": "bp_explore_beach", "interactionType": "EXPLORE", "eventId": None, "functionType": None,
     "subSceneId": None, "direction": None, "targetSceneId": None, "targetSubSceneId": None,
     "targetMapId": None, "targetNodeId": None, "travelTimeMinutes": None, "staminaCost": None,
     "pathDescription": None, "encounterEventPool": None, "requirements": None, "traderId": None},
    {"parent_id": "bp_enter_cave", "interactionType": "ENTER_SUB_SCENE", "eventId": None, "functionType": None,
     "subSceneId": "beach_cave", "direction": None, "targetSceneId": None, "targetSubSceneId": None,
     "targetMapId": None, "targetNodeId": None, "travelTimeMinutes": None, "staminaCost": None,
     "pathDescription": None, "encounterEventPool": None, "requirements": None, "traderId": None},
    {"parent_id": "bp_move_to_forest", "interactionType": "MOVE_TO_SCENE", "eventId": None, "functionType": None,
     "subSceneId": None, "direction": None, "targetSceneId": "forest", "targetSubSceneId": None,
     "targetMapId": None, "targetNodeId": "node_forest", "travelTimeMinutes": 30, "staminaCost": 25,
     "pathDescription": "沿着沙滩向北走，进入森林",
     "encounterEventPool": "event_forest_path_encounter:30,event_merchant_encounter:10",
     "requirements": None, "traderId": None},
    {"parent_id": "bp_rest_beach", "interactionType": "REST", "eventId": None, "functionType": None,
     "subSceneId": None, "direction": None, "targetSceneId": None, "targetSubSceneId": None,
     "targetMapId": None, "targetNodeId": None, "travelTimeMinutes": None, "staminaCost": None,
     "pathDescription": None, "encounterEventPool": None, "requirements": None, "traderId": None},
    # 森林行为参数
    {"parent_id": "bp_explore_forest", "interactionType": "EXPLORE", "eventId": None, "functionType": None,
     "subSceneId": None, "direction": None, "targetSceneId": None, "targetSubSceneId": None,
     "targetMapId": None, "targetNodeId": None, "travelTimeMinutes": None, "staminaCost": None,
     "pathDescription": None, "encounterEventPool": None, "requirements": None, "traderId": None},
    {"parent_id": "bp_return_to_beach", "interactionType": "MOVE_TO_SCENE", "eventId": None, "functionType": None,
     "subSceneId": None, "direction": None, "targetSceneId": "beach", "targetSubSceneId": None,
     "targetMapId": None, "targetNodeId": "node_beach", "travelTimeMinutes": 30, "staminaCost": 25,
     "pathDescription": "穿过林间小径返回海滩",
     "encounterEventPool": None, "requirements": None, "traderId": None},
    {"parent_id": "bp_rest_forest", "interactionType": "REST", "eventId": None, "functionType": None,
     "subSceneId": None, "direction": None, "targetSceneId": None, "targetSubSceneId": None,
     "targetMapId": None, "targetNodeId": None, "travelTimeMinutes": None, "staminaCost": None,
     "pathDescription": None, "encounterEventPool": None, "requirements": None, "traderId": None},
    # 洞穴行为参数
    {"parent_id": "bp_explore_cave", "interactionType": "EXPLORE", "eventId": None, "functionType": None,
     "subSceneId": None, "direction": None, "targetSceneId": None, "targetSubSceneId": None,
     "targetMapId": None, "targetNodeId": None, "travelTimeMinutes": None, "staminaCost": None,
     "pathDescription": None, "encounterEventPool": None, "requirements": None, "traderId": None},
    {"parent_id": "bp_exit_cave", "interactionType": "EXIT_SUB_SCENE", "eventId": None, "functionType": None,
     "subSceneId": None, "direction": None, "targetSceneId": None, "targetSubSceneId": None,
     "targetMapId": None, "targetNodeId": None, "travelTimeMinutes": None, "staminaCost": None,
     "pathDescription": None, "encounterEventPool": None, "requirements": None, "traderId": None},
]

for row_idx, bp in enumerate(behavior_params, 2):
    write_row(ws_bp, row_idx, [
        bp["parent_id"], bp["interactionType"], bp["eventId"], bp["functionType"],
        bp["subSceneId"], bp["direction"], bp["targetSceneId"], bp["targetSubSceneId"],
        bp["targetMapId"], bp["targetNodeId"], bp["travelTimeMinutes"], bp["staminaCost"],
        bp["pathDescription"], bp["encounterEventPool"], bp["requirements"], bp["traderId"]
    ])

ws_bp.column_dimensions['A'].width = 24
ws_bp.column_dimensions['B'].width = 18
ws_bp.column_dimensions['C'].width = 20
ws_bp.column_dimensions['D'].width = 16
ws_bp.column_dimensions['E'].width = 16
ws_bp.column_dimensions['F'].width = 12
ws_bp.column_dimensions['G'].width = 16
ws_bp.column_dimensions['H'].width = 16
ws_bp.column_dimensions['I'].width = 16
ws_bp.column_dimensions['J'].width = 16
ws_bp.column_dimensions['K'].width = 16
ws_bp.column_dimensions['L'].width = 14
ws_bp.column_dimensions['M'].width = 30
ws_bp.column_dimensions['N'].width = 40
ws_bp.column_dimensions['O'].width = 30
ws_bp.column_dimensions['P'].width = 16


# ============================================================
# 创建条件.xlsx
# ============================================================

cond_wb = openpyxl.Workbook()

# ---- 条件 Sheet ----
ws_cond = cond_wb.active
ws_cond.title = "条件"
write_header(ws_cond, [
    "id", "logic\n逻辑运算符", "subConditions\n子条件列表(逗号分隔)",
    "target\n条件目标(条件目标id)", "operator\n比较运算符",
    "value", "value2"
])

conditions = [
    # 复合条件
    {"id": "cond_beach_crabs", "logic": "AND", "subConditions": "cond_first_time_on_beach,cond_corruption_25",
     "target": None, "operator": None, "value": None, "value2": None},
    {"id": "cond_cave_first_enter", "logic": "AND", "subConditions": "cond_explored_cave_false",
     "target": None, "operator": None, "value": None, "value2": None},
    {"id": "cond_enter_cave", "logic": "AND", "subConditions": "cond_explored_cave_true",
     "target": None, "operator": None, "value": None, "value2": None},
    # 简单条件
    {"id": "cond_first_time_on_beach", "logic": None, "subConditions": None,
     "target": "tgt_first_time_on_beach", "operator": "EQUAL", "value": "true", "value2": None},
    {"id": "cond_corruption_25", "logic": None, "subConditions": None,
     "target": "tgt_corruption", "operator": "GREATER_EQUAL", "value": 25, "value2": None},
    {"id": "cond_explored_cave_false", "logic": None, "subConditions": None,
     "target": "tgt_explored_cave", "operator": "EQUAL", "value": "false", "value2": None},
    {"id": "cond_explored_cave_true", "logic": None, "subConditions": None,
     "target": "tgt_explored_cave", "operator": "NOT_EQUAL", "value": "false", "value2": None},
    {"id": "cond_strange_markings", "logic": None, "subConditions": None,
     "target": "tgt_san_60", "operator": "LESS_EQUAL", "value": 60, "value2": None},
]

for row_idx, cond in enumerate(conditions, 2):
    write_row(ws_cond, row_idx, [
        cond["id"], cond["logic"], cond["subConditions"],
        cond["target"], cond["operator"], cond["value"], cond["value2"]
    ])

ws_cond.column_dimensions['A'].width = 24
ws_cond.column_dimensions['B'].width = 14
ws_cond.column_dimensions['C'].width = 30
ws_cond.column_dimensions['D'].width = 24
ws_cond.column_dimensions['E'].width = 18
ws_cond.column_dimensions['F'].width = 14
ws_cond.column_dimensions['G'].width = 10

# ---- 条件目标 Sheet ----
ws_target = cond_wb.create_sheet("条件目标")
write_header(ws_target, [
    "thisID(条件目标id)", "type", "id", "attributeType", "subType"
])

targets = [
    {"thisID": "tgt_first_time_on_beach", "type": "FLAG", "id": "first_time_on_beach", "attributeType": None, "subType": None},
    {"thisID": "tgt_corruption", "type": "CORRUPTION", "id": None, "attributeType": None, "subType": None},
    {"thisID": "tgt_explored_cave", "type": "FLAG", "id": "explored_cave", "attributeType": None, "subType": None},
    {"thisID": "tgt_san_60", "type": "ATTRIBUTE", "id": None, "attributeType": "SAN", "subType": None},
]

for row_idx, target in enumerate(targets, 2):
    write_row(ws_target, row_idx, [
        target["thisID"], target["type"], target["id"],
        target["attributeType"], target["subType"]
    ])

ws_target.column_dimensions['A'].width = 24
ws_target.column_dimensions['B'].width = 16
ws_target.column_dimensions['C'].width = 24
ws_target.column_dimensions['D'].width = 20
ws_target.column_dimensions['E'].width = 14


# ============================================================
# 保存文件
# ============================================================

scene_wb.save(r"d:\WORK\RPGame-demo\剧本设计\场景.xlsx")
cond_wb.save(r"d:\WORK\RPGame-demo\剧本设计\条件.xlsx")
print("场景.xlsx 和 条件.xlsx 已成功生成！")